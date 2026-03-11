#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import json
import math
import argparse
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from dataclasses import dataclass
from typing import List, Tuple, Optional

def centered_slope(y: np.ndarray, dt: float, w: int = 3) -> np.ndarray:
    y = np.asarray(y, dtype=float)
    dy = np.full_like(y, np.nan)
    for i in range(w, len(y) - w):
        dy[i] = (y[i + w] - y[i - w]) / (2 * w * dt)
    return dy

@dataclass
class Rep:
    idx_trough: int
    idx_peak: int
    t_trough: float
    t_peak: float
    amp_mm: float
    ok_85: int
    dur_s: float
    v_max_mms: float
    v_mean_mms: float
    p_max_w: Optional[float]
    p_mean_w: Optional[float]
    work_j: Optional[float]

def detect_pairs(z: np.ndarray, vel: np.ndarray, dt: float,
                 min_separation_s: float, min_prom_frac: float) -> List[Tuple[int, int]]:
    N = len(z)
    sign = np.sign(vel)
    sign[np.isnan(sign)] = 0
    zc_idx = np.where(np.diff(sign) != 0)[0] + 1
    troughs, peaks = [], []
    for idx in zc_idx:
        prev = vel[idx - 1] if idx - 1 >= 0 else np.nan
        cur = vel[idx]
        if np.isnan(prev) or np.isnan(cur):
            continue
        i0, i1 = max(0, idx - 3), min(N, idx + 4)
        if prev < 0 and cur > 0:
            troughs.append(np.argmin(z[i0:i1]) + i0)
        elif prev > 0 and cur < 0:
            peaks.append(np.argmax(z[i0:i1]) + i0)
    troughs = sorted(set(troughs))
    peaks = sorted(set(peaks))

    min_sep = int(round(min_separation_s / dt))
    z_range = float(np.nanmax(z) - np.nanmin(z))

    rep_pairs = []
    pi = 0
    last_p = -10**9
    for ti in troughs:
        while pi < len(peaks) and peaks[pi] <= ti:
            pi += 1
        if pi >= len(peaks):
            break
        p = peaks[pi]
        if p - last_p < min_sep:
            if rep_pairs:
                prev_t, prev_p = rep_pairs[-1]
                prev_amp = z[prev_p] - z[prev_t]
                new_amp = z[p] - z[ti]
                if new_amp > prev_amp:
                    rep_pairs[-1] = (ti, p)
                    last_p = p
            continue
        if (z[p] - z[ti]) >= (min_prom_frac * z_range):
            rep_pairs.append((ti, p))
            last_p = p
            pi += 1
        else:
            pi += 1
    return rep_pairs

def auto_tune_detection(z: np.ndarray, vel: np.ndarray, dt: float,
                        target_range=(8, 20)) -> Tuple[List[Tuple[int,int]], float, float]:
    cand_min_sep = [0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8]
    cand_prom = [0.005, 0.01, 0.015, 0.02, 0.03, 0.04]
    best = None
    best_score = -1e9
    z_range = float(np.nanmax(z) - np.nanmin(z))
    for ms in cand_min_sep:
        for pm in cand_prom:
            pairs = detect_pairs(z, vel, dt, ms, pm)
            n = len(pairs)
            amps = [z[p] - z[t0] for (t0, p) in pairs]
            med_amp = float(np.nanmedian(amps)) if amps else 0.0
            penalty = 0.0
            if n < target_range[0]:
                penalty -= (target_range[0] - n) * 2.0
            elif n > target_range[1]:
                penalty -= (n - target_range[1]) * 2.0
            score = -abs(n - 13) + 0.2 * (med_amp / (z_range + 1e-9)) + penalty
            if score > best_score:
                best_score = score
                best = (pairs, ms, pm)
    if best is None:
        pairs = detect_pairs(z, vel, dt, 0.75, 0.02)
        return pairs, 0.75, 0.02
    return best

def compute_power_work(time_s: np.ndarray, v_m_s: np.ndarray,
                       Fz_N: Optional[np.ndarray] = None, mass_kg: Optional[float] = None):
    if Fz_N is not None:
        P = Fz_N * v_m_s
        return P, 'P = Fz * v_z'
    elif mass_kg is not None:
        g = 9.80665
        P = mass_kg * g * v_m_s
        return P, 'P = m * g * v_z (aprox.)'
    else:
        return np.full_like(v_m_s, np.nan), 'No se encontró Fz ni masa: P no disponible'

def run_analysis(file_path: str,
                 out_prefix: Optional[str] = None,
                 sheet_name: str = 'Reducido',
                 time_col: str = 'time',
                 bcm_z_col: Optional[str] = None,
                 start_col_excel: str = 'CW',
                 mass_kg: Optional[float] = None,
                 grf_sheet: Optional[str] = 'Plataformas de fuerzas',
                 grf_time_col: Optional[str] = 'time',
                 grf_z_col: Optional[str] = None,
                 min_sep_s: Optional[float] = None,
                 prom_frac: Optional[float] = None,
                 make_plot: bool = True,
                 ):
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    if time_col not in df.columns:
        raise ValueError(f'No se encontró la columna de tiempo \"{time_col}\" en la hoja {sheet_name}.')
    t = df[time_col].to_numpy(dtype=float)
    dt = float(pd.Series(np.diff(t)).median())

    if bcm_z_col is None:
        cand_cols = [c for c in df.columns
                     if 'Body center of mass' in c
                     and ('-z' in c.lower() or ' z ' in c.lower())
                     and ('(mm)' in c.lower() or '[mm]' in c.lower())]
        if not cand_cols:
            cand_cols = [c for c in df.columns if 'Body center of mass' in c and ('z' in c.lower())]
        if not cand_cols:
            raise ValueError('No se encontró columna de desplazamiento BCM en Z (mm).')
        bcm_z_col = cand_cols[0]

    z_mm = df[bcm_z_col].to_numpy(dtype=float)
    z_m = z_mm / 1000.0

    vel_m_s = centered_slope(z_m, dt, w=3)
    acc_m_s2 = centered_slope(vel_m_s, dt, w=3)
    vel_mm_s = vel_m_s * 1000.0

    # Cargar Fz si está disponible
    Fz = None
    grf_info = None
    if grf_sheet is not None:
        try:
            dfg = pd.read_excel(file_path, sheet_name=grf_sheet, engine='openpyxl')
            if grf_z_col is None:
                grf_candidates = [c for c in dfg.columns
                                  if ('fz' in c.lower() or 'force z' in c.lower() or 'vertical' in c.lower())
                                  and ('n' in c.lower() or ' (n' in c.lower())]
                if not grf_candidates:
                    grf_candidates = [c for c in dfg.columns if ('force' in c.lower() and 'z' in c.lower())]
                if grf_candidates:
                    grf_z_col = grf_candidates[0]
            if grf_z_col and (grf_time_col in dfg.columns):
                dfg2 = dfg[[grf_time_col, grf_z_col]].dropna().sort_values(grf_time_col)
                df_main = pd.DataFrame({time_col: t})
                merged = pd.merge_asof(df_main, dfg2, left_on=time_col, right_on=grf_time_col,
                                       direction='nearest', tolerance=None)
                if grf_time_col in merged.columns:
                    merged = merged.drop(columns=[grf_time_col])
                Fz = merged[grf_z_col].to_numpy(dtype=float)
                grf_info = f'GRFz: hoja \"{grf_sheet}\", columna \"{grf_z_col}\"'
        except Exception:
            pass

    # Potencia y trabajo a nivel de señal completa
    P_w, p_model = compute_power_work(t, vel_m_s, Fz_N=Fz, mass_kg=mass_kg)

    # Detección (auto-tune salvo que se fijen parámetros)
    if (min_sep_s is not None) and (prom_frac is not None):
        pairs = detect_pairs(z_mm, vel_mm_s, dt, min_sep_s, prom_frac)
        used_min_sep, used_prom = min_sep_s, prom_frac
    else:
        pairs, used_min_sep, used_prom = auto_tune_detection(z_mm, vel_mm_s, dt)

    amps_mm = [z_mm[p] - z_mm[t0] for (t0, p) in pairs]
    max_amp = float(np.nanmax(amps_mm)) if amps_mm else np.nan
    ok_flags = [(1 if (a >= 0.85 * max_amp) else 0) for a in amps_mm] if amps_mm else []

    # Anotaciones por muestra
    N = len(z_mm)
    evt_trough = np.zeros(N, dtype=int)
    evt_peak = np.zeros(N, dtype=int)
    rep_idx = np.full(N, np.nan)
    rep_amp = np.full(N, np.nan)
    rep_ok = np.full(N, np.nan)

    reps: List[Rep] = []
    for k, (t0, p) in enumerate(pairs, start=1):
        evt_trough[t0] = 1
        evt_peak[p] = 1
        rep_idx[t0:p + 1] = k
        amp = z_mm[p] - z_mm[t0]
        rep_amp[t0:p + 1] = amp
        rep_ok[t0:p + 1] = ok_flags[k - 1]
        seg_vel = vel_mm_s[t0:p + 1]
        seg_P = P_w[t0:p + 1] if not np.all(np.isnan(P_w)) else None
        v_max = float(np.nanmax(seg_vel)) if seg_vel.size else np.nan
        v_mean = float(np.nanmean(seg_vel[seg_vel > 0])) if np.any(seg_vel > 0) else float(np.nanmean(seg_vel))
        dur = float((p - t0) * dt)
        if seg_P is not None:
            p_max = float(np.nanmax(seg_P))
            p_mean = float(np.nanmean(seg_P))
            work = float(np.trapezoid(seg_P, dx=dt))
        else:
            p_max = p_mean = work = None
        reps.append(Rep(t0, p, float(t[t0]), float(t[p]), float(amp), int(ok_flags[k - 1]),
                        dur, v_max, v_mean, p_max, p_mean, work))

    # Salidas
    out_prefix = out_prefix or (file_path.rsplit('.xlsx', 1)[0] + '_analysis')
    out_excel = out_prefix + '.xlsx'
    out_plot = out_prefix + '_segmentation.png'
    out_params = out_prefix + '_params.json'

    # Escribir al Excel fuente + hojas nuevas
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        start_col_idx = column_index_from_string(start_col_excel)
        start_row = 1
        out_cols = pd.DataFrame({
            'BCM_z_mm (orig)': z_mm,
            'BCM_z_vel_c7 (mm/s)': vel_mm_s,
            'BCM_z_acc_c7 (mm/s^2)': acc_m_s2 * 1000.0,
            'Power (W)': P_w,
            'evt_trough': evt_trough,
            'evt_peak': evt_peak,
            'rep_index_trough_to_peak': rep_idx,
            'rep_amp_mm': rep_amp,
            'rep_ok_85pct': rep_ok,
        })
        for j, col in enumerate(out_cols.columns, start=start_col_idx):
            ws.cell(row=start_row, column=j, value=col)
        for i in range(N):
            for j, col in enumerate(out_cols.columns, start=start_col_idx):
                val = out_cols.iloc[i, j - start_col_idx]
                if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                    val = None
                ws.cell(row=start_row + 1 + i, column=j, value=val)

        if 'STS Events (BCM Z)' in wb.sheetnames:
            del wb['STS Events (BCM Z)']
        ws_ev = wb.create_sheet('STS Events (BCM Z)')
        ev_cols = ['rep', 'trough_idx', 'peak_idx', 'trough_time_s', 'peak_time_s',
                   'amplitude_mm', 'ok_85pct', 'duration_s', 'vmax_mm_s', 'vmean_mm_s',
                   'pmax_W', 'pmean_W', 'work_J']
        for j, col in enumerate(ev_cols, start=1):
            ws_ev.cell(row=1, column=j, value=col)
        for i, r in enumerate(reps, start=2):
            row = [i - 1, r.idx_trough, r.idx_peak, r.t_trough, r.t_peak, r.amp_mm, r.ok_85,
                   r.dur_s, r.v_max_mms, r.v_mean_mms, r.p_max_w, r.p_mean_w, r.work_j]
            for j, val in enumerate(row, start=1):
                ws_ev.cell(row=i, column=j, value=val)

        if 'STS Summary' in wb.sheetnames:
            del wb['STS Summary']
        ws_sm = wb.create_sheet('STS Summary')
        n_reps = len(reps)
        n_ok = int(sum(r.ok_85 for r in reps))
        amp_vals = np.array([r.amp_mm for r in reps], float)
        dur_vals = np.array([r.dur_s for r in reps], float)
        vmax_vals = np.array([r.v_max_mms for r in reps], float)
        vmean_vals = np.array([r.v_mean_mms for r in reps], float)

        summary = {
            'n_reps_total': n_reps,
            'n_reps_ok_85pct': n_ok,
            'max_amp_mm': float(np.nanmax(amp_vals)) if n_reps else None,
            'amp_mean_mm': float(np.nanmean(amp_vals)) if n_reps else None,
            'amp_sd_mm': float(np.nanstd(amp_vals, ddof=1)) if n_reps > 1 else None,
            'amp_cv_pct': float(np.nanstd(amp_vals, ddof=1) / np.nanmean(amp_vals) * 100.0) if n_reps > 1 else None,
            'dur_mean_s': float(np.nanmean(dur_vals)) if n_reps else None,
            'dur_sd_s': float(np.nanstd(dur_vals, ddof=1)) if n_reps > 1 else None,
            'vmax_mean_mm_s': float(np.nanmean(vmax_vals)) if n_reps else None,
            'vmax_sd_mm_s': float(np.nanstd(vmax_vals, ddof=1)) if n_reps > 1 else None,
            'vmean_mean_mm_s': float(np.nanmean(vmean_vals)) if n_reps else None,
            'vmean_sd_mm_s': float(np.nanstd(vmean_vals, ddof=1)) if n_reps > 1 else None,
            'power_model': p_model,
            'params_min_sep_s': used_min_sep,
            'params_prom_frac': used_prom,
            'bcm_z_col': bcm_z_col,
            'grf_info': grf_info,
            'mass_kg': mass_kg,
        }
        r = 1
        for k, v in summary.items():
            ws_sm.cell(row=r, column=1, value=k)
            ws_sm.cell(row=r, column=2, value=v)
            r += 1

        wb.save(out_excel)
    except Exception as e:
        # Fallback: crear un Excel nuevo con todo
        events_df = pd.DataFrame([{k: getattr(r, k) for k in r.__dataclass_fields__.keys()} for r in reps])
        with pd.ExcelWriter(out_excel, engine='openpyxl') as xw:
            df_out = df.copy()
            df_out['BCM_z_mm (orig)'] = z_mm
            df_out['BCM_z_vel_c7 (mm/s)'] = vel_mm_s
            df_out['BCM_z_acc_c7 (mm/s^2)'] = acc_m_s2 * 1000.0
            df_out['Power (W)'] = P_w
            df_out['evt_trough'] = evt_trough
            df_out['evt_peak'] = evt_peak
            df_out['rep_index_trough_to_peak'] = rep_idx
            df_out['rep_amp_mm'] = rep_amp
            df_out['rep_ok_85pct'] = rep_ok
            df_out.to_excel(xw, sheet_name=sheet_name, index=False)
            events_df.to_excel(xw, sheet_name='STS Events (BCM Z)', index=False)
            pd.DataFrame({'key': ['error'], 'value': [str(e)]}).to_excel(xw, sheet_name='STS Summary', index=False)

    # Gráfico
    if make_plot:
        fig, ax1 = plt.subplots(figsize=(10, 6))
        ax2 = ax1.twinx()
        ax1.plot(t, z_mm, color='tab:blue', label='Desplazamiento BCM Z (mm)')
        ax2.plot(t, vel_mm_s, color='tab:orange', alpha=0.7, label='Velocidad BCM Z (mm/s)')
        for k, (t0, p) in enumerate(pairs, start=1):
            ax1.axvline(t[t0], color='royalblue', linestyle='--', alpha=0.6)
            ax1.axvline(t[p], color='crimson', linestyle='--', alpha=0.6)
            ax1.text(t[p], z_mm[p], f'{k}', color='crimson', fontsize=8, ha='left', va='bottom')
            if ok_flags[k - 1] == 1:
                ax1.fill_between([t[t0], t[p]], [z_mm[t0], z_mm[t0]], [z_mm[p], z_mm[p]], color='green', alpha=0.1)
        ax1.set_xlabel('Tiempo (s)')
        ax1.set_ylabel('Desplazamiento (mm)')
        ax2.set_ylabel('Velocidad (mm/s)')
        ax1.set_title('Segmentación STS con BCM Z')
        l1, _ = ax1.get_legend_handles_labels()
        l2, _ = ax2.get_legend_handles_labels()
        ax1.legend(l1 + l2, [h.get_label() for h in l1 + l2], loc='upper left')
        plt.tight_layout()
        fig.savefig(out_plot, dpi=150)
        plt.close(fig)

    # Parámetros y meta
    params = {
        'file_path': file_path,
        'sheet_name': sheet_name,
        'time_col': time_col,
        'bcm_z_col': bcm_z_col,
        'dt': dt,
        'min_sep_s': used_min_sep,
        'prom_frac': used_prom,
        'n_reps': len(pairs),
        'n_ok_85pct': int(sum(ok_flags)) if ok_flags else 0,
        'power_model': p_model,
        'mass_kg': mass_kg,
        'grf_sheet': grf_sheet,
        'grf_z_col': grf_z_col,
    }
    with open(out_params, 'w', encoding='utf-8') as f:
        json.dump(params, f, ensure_ascii=False, indent=2)

    print('\\nLISTO')
    print('Excel:', out_excel)
    print('Gráfico:', out_plot)
    print('Parámetros:', out_params)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Herramienta de análisis STS con BCM Z')
    parser.add_argument('--input', required=True, help='Ruta del archivo .xlsx de entrada')
    parser.add_argument('--sheet', default='Reducido', help='Hoja a procesar (por defecto: Reducido)')
    parser.add_argument('--time-col', default='time', help='Nombre de la columna de tiempo (por defecto: time)')
    parser.add_argument('--bcm-z-col', default=None, help='Nombre exacto de la columna de BCM Z (mm); si no, se autodetecta')
    parser.add_argument('--start-col', default='CW', help='Columna inicial para escribir resultados en Excel (por defecto: CW)')
    parser.add_argument('--mass-kg', type=float, default=None,
                        help='Masa corporal (kg) para potencia P=m*g*v si no hay Fz')
    parser.add_argument('--grf-sheet', default='Plataformas de fuerzas',
                        help='Hoja con fuerzas (por defecto: Plataformas de fuerzas)')
    parser.add_argument('--grf-time-col', default='time', help='Columna de tiempo en hoja de fuerzas (por defecto: time)')
    parser.add_argument('--grf-z-col', default=None, help='Columna de Fz (N); si no, se intenta autodetectar')
    parser.add_argument('--min-sep', type=float, default=None, help='Separación mínima entre reps (s). Si no, auto-tune')
    parser.add_argument('--prom-frac', type=float, default=None, help='Prominencia mínima relativa (0-1). Si no, auto-tune')
    parser.add_argument('--no-plot', action='store_true', help='No generar gráfico')
    parser.add_argument('--out-prefix', default=None, help='Prefijo de salida (sin extensión)')

    args = parser.parse_args()

    run_analysis(file_path=args.input,
                 out_prefix=args.out_prefix,
                 sheet_name=args.sheet,
                 time_col=args.time_col,
                 bcm_z_col=args.bcm_z_col,
                 start_col_excel=args.start_col,
                 mass_kg=args.mass_kg,
                 grf_sheet=args.grf_sheet,
                 grf_time_col=args.grf_time_col,
                 grf_z_col=args.grf_z_col,
                 min_sep_s=args.min_sep,
                 prom_frac=args.prom_frac,
                 make_plot=(not args.no_plot))