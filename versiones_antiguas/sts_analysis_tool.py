#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Herramienta de análisis STS (sit-to-stand) basada en BCM Z.
Fases por ciclo: Levantarse (valle->pico), Sentarse (pico->valle), Sentado.
Calcula velocidad/aceleración (derivada centrada 7), potencia/trabajo por masa (P=m*g*v),
segmenta automáticamente o con parámetros fijos, genera métricas por repetición y globales,
y crea gráficos general y por repetición.
"""

import argparse
import json
import math
import os
from dataclasses import dataclass
from typing import List, Tuple, Optional

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# ---------------- Helpers ----------------

def centered_slope(y: np.ndarray, dt: float, w: int = 3) -> np.ndarray:
    y = np.asarray(y, dtype=float)
    dy = np.full_like(y, np.nan)
    for i in range(w, len(y) - w):
        dy[i] = (y[i + w] - y[i - w]) / (2 * w * dt)
    return dy


@dataclass
class RepUP:  # Levantarse (valle -> pico)
    idx_trough: int
    idx_peak: int
    t_trough: float
    t_peak: float
    amp_mm: float
    ok: int
    dur_s: float
    v_max_mms: float
    v_mean_mms: float
    t_to_vmax_s: float
    t_pos_acc_s: float
    p_max_w: Optional[float]
    p_mean_w: Optional[float]
    work_j: Optional[float]


@dataclass
class RepDOWN:  # Sentarse (pico -> valle)
    idx_peak: int
    idx_trough_next: int
    t_peak: float
    t_trough_next: float
    amp_mm: float  # caída en mm (pico->valle siguiente)
    dur_s: float
    v_min_mms: float
    v_mean_mms: float
    t_to_vmin_s: float
    t_neg_acc_s: float
    p_min_w: Optional[float]
    p_mean_w: Optional[float]
    work_j: Optional[float]


@dataclass
class PhaseSEATED:  # Sentado (valle_i+1 -> inicio siguiente levantarse)
    idx_start: int
    idx_end: int
    t_start: float
    t_end: float
    dur_s: float
    v_abs_mean_mms: float
    a_abs_mean_mms2: float


# ---------------- Detección ----------------

def detect_pairs(z: np.ndarray, vel: np.ndarray, dt: float,
                 min_separation_s: float, min_prom_frac: float) -> List[Tuple[int, int]]:
    """Detecta pares valle->pico con refinamiento local en ±3 muestras."""
    N = len(z)
    sign = np.sign(vel)
    sign[np.isnan(sign)] = 0
    zc_idx = np.where(np.diff(sign) != 0)[0] + 1  # cruces por cero
    troughs, peaks = [], []
    for idx in zc_idx:
        prev = vel[idx - 1] if idx - 1 >= 0 else np.nan
        cur = vel[idx]
        if np.isnan(prev) or np.isnan(cur):
            continue
        i0, i1 = max(0, idx - 3), min(N, idx + 4)
        if prev < 0 and cur > 0:
            troughs.append(np.argmin(z[i0:i1]) + i0)
        elif prev > 0 and cur < 0:
            peaks.append(np.argmax(z[i0:i1]) + i0)
    troughs = sorted(set(troughs))
    peaks = sorted(set(peaks))

    min_sep = int(round(min_separation_s / dt))
    z_range = float(np.nanmax(z) - np.nanmin(z))

    rep_pairs = []
    pi = 0
    last_p = -10**9
    for ti in troughs:
        while pi < len(peaks) and peaks[pi] <= ti:
            pi += 1
        if pi >= len(peaks):
            break
        p = peaks[pi]
        if p - last_p < min_sep:
            if rep_pairs:
                prev_t, prev_p = rep_pairs[-1]
                prev_amp = z[prev_p] - z[prev_t]
                new_amp = z[p] - z[ti]
                if new_amp > prev_amp:
                    rep_pairs[-1] = (ti, p)
                    last_p = p
            continue
        if (z[p] - z[ti]) >= (min_prom_frac * z_range):
            rep_pairs.append((ti, p))
            last_p = p
            pi += 1
        else:
            pi += 1
    return rep_pairs


def auto_tune_detection(z: np.ndarray, vel: np.ndarray, dt: float,
                        target_range=(8, 20)) -> Tuple[List[Tuple[int, int]], float, float]:
    """Búsqueda gruesa de (min_sep, prom_frac) para estabilizar el conteo de repeticiones."""
    cand_min_sep = [0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8]
    cand_prom = [0.005, 0.01, 0.015, 0.02, 0.03, 0.04]
    best = None
    best_score = -1e9
    z_range = float(np.nanmax(z) - np.nanmin(z))
    for ms in cand_min_sep:
        for pm in cand_prom:
            pairs = detect_pairs(z, vel, dt, ms, pm)
            n = len(pairs)
            amps = [z[p] - z[t0] for (t0, p) in pairs]
            med_amp = float(np.nanmedian(amps)) if amps else 0.0
            penalty = 0.0
            if n < target_range[0]:
                penalty -= (target_range[0] - n) * 2.0
            elif n > target_range[1]:
                penalty -= (n - target_range[1]) * 2.0
            score = -abs(n - 13) + 0.2 * (med_amp / (z_range + 1e-9)) + penalty
            if score > best_score:
                best_score = score
                best = (pairs, ms, pm)
    if best is None:
        pairs = detect_pairs(z, vel, dt, 0.75, 0.02)
        return pairs, 0.75, 0.02
    return best


# ---------------- Potencia/Trabajo ----------------

def compute_power_mass(vel_m_s: np.ndarray, mass_kg: Optional[float]) -> Tuple[np.ndarray, str]:
    if mass_kg is None:
        return np.full_like(vel_m_s, np.nan), 'Potencia no disponible (sin masa)'
    g = 9.80665
    return mass_kg * g * vel_m_s, 'P = m * g * v_z (aprox.)'


# ---------------- Núcleo de la herramienta ----------------

def run_tool(
    file_path: str,
    out_prefix: Optional[str] = None,
    sheet_name: str = 'Reducido',
    time_col: str = 'time',
    bcm_z_col: Optional[str] = None,
    start_col_excel: str = 'CW',
    mass_kg: Optional[float] = None,
    min_sep_s: Optional[float] = None,
    prom_frac: Optional[float] = None,
    ok_th: float = 0.85,
    make_plot: bool = True,
    per_rep_plots: bool = True,
):
    # Cargar hoja
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    if time_col not in df.columns:
        raise ValueError(f'No se encontró la columna de tiempo "{time_col}" en la hoja {sheet_name}.')
    t = df[time_col].to_numpy(dtype=float)
    dt = float(pd.Series(np.diff(t)).median())

    # Columna BCM Z
    if bcm_z_col is None:
        cand_cols = [c for c in df.columns if 'Body center of mass' in c and ('-z' in c.lower() or ' z ' in c.lower()) and ('(mm)' in c.lower() or '[mm]' in c.lower())]
        if not cand_cols:
            cand_cols = [c for c in df.columns if 'Body center of mass' in c and ('z' in c.lower())]
        if not cand_cols:
            raise ValueError('No se encontró columna de desplazamiento BCM en Z (mm).')
        bcm_z_col = cand_cols[0]

    z_mm = df[bcm_z_col].to_numpy(dtype=float)
    z_m = z_mm / 1000.0

    # Derivadas centradas 7
    vel_m_s = centered_slope(z_m, dt, w=3)
    acc_m_s2 = centered_slope(vel_m_s, dt, w=3)
    vel_mm_s = vel_m_s * 1000.0

    # Potencia por masa
    P_w, p_model = compute_power_mass(vel_m_s, mass_kg)

    # Detección UP (valle->pico)
    if (min_sep_s is not None) and (prom_frac is not None):
        up_pairs = detect_pairs(z_mm, vel_mm_s, dt, min_sep_s, prom_frac)
        used_min_sep, used_prom = min_sep_s, prom_frac
    else:
        up_pairs, used_min_sep, used_prom = auto_tune_detection(z_mm, vel_mm_s, dt)

    amps_up = [z_mm[p] - z_mm[t0] for (t0, p) in up_pairs]
    max_amp_up = float(np.nanmax(amps_up)) if amps_up else np.nan
    ok_flags_up = [(1 if (a >= ok_th * max_amp_up) else 0) for a in amps_up] if amps_up else []

    N = len(z_mm)
    evt_trough = np.zeros(N, dtype=int)
    evt_peak = np.zeros(N, dtype=int)

    reps_up: List[RepUP] = []
    for k, (t0, p) in enumerate(up_pairs, start=1):
        evt_trough[t0] = 1
        evt_peak[p] = 1
        seg_v = vel_mm_s[t0:p+1]
        seg_acc = acc_m_s2[t0:p+1] * 1000.0
        seg_P = P_w[t0:p+1] if not np.all(np.isnan(P_w)) else None
        v_max = float(np.nanmax(seg_v)) if seg_v.size else np.nan
        v_mean = float(np.nanmean(seg_v[seg_v > 0])) if np.any(seg_v > 0) else float(np.nanmean(seg_v))
        imax = np.nanargmax(seg_v) if np.any(~np.isnan(seg_v)) else 0
        t_to_vmax = imax * dt
        t_pos_acc = float(np.sum((seg_acc > 0).astype(int)) * dt)
        dur = float((p - t0) * dt)
        if seg_P is not None:
            p_max = float(np.nanmax(seg_P))
            p_mean = float(np.nanmean(seg_P))
            work = float(np.trapezoid(seg_P, dx=dt))
        else:
            p_max = p_mean = work = None
        reps_up.append(RepUP(t0, p, float(t[t0]), float(t[p]), float(z_mm[p]-z_mm[t0]), int(ok_flags_up[k-1]) if ok_flags_up else 0,
                             dur, v_max, v_mean, t_to_vmax, t_pos_acc, p_max, p_mean, work))

    # Construir DOWN (pico->valle siguiente) y SEATED
    troughs_sorted = [ru.idx_trough for ru in reps_up]
    peaks_sorted = [ru.idx_peak for ru in reps_up]
    reps_down: List[RepDOWN] = []
    phases_seated: List[PhaseSEATED] = []
    for i in range(min(len(peaks_sorted), len(troughs_sorted)-1)):
        p = peaks_sorted[i]
        nv = troughs_sorted[i+1]
        # DOWN
        seg_v = vel_mm_s[p:nv+1]
        seg_acc = acc_m_s2[p:nv+1] * 1000.0
        seg_P = P_w[p:nv+1] if not np.all(np.isnan(P_w)) else None
        v_min = float(np.nanmin(seg_v)) if seg_v.size else np.nan
        v_mean = float(np.nanmean(seg_v[seg_v < 0])) if np.any(seg_v < 0) else float(np.nanmean(seg_v))
        imin = np.nanargmin(seg_v) if np.any(~np.isnan(seg_v)) else 0
        t_to_vmin = imin * dt
        t_neg_acc = float(np.sum((seg_acc < 0).astype(int)) * dt)
        dur = float((nv - p) * dt)
        if seg_P is not None:
            p_min = float(np.nanmin(seg_P))
            p_mean = float(np.nanmean(seg_P))
            work = float(np.trapezoid(seg_P, dx=dt))
        else:
            p_min = p_mean = work = None
        reps_down.append(RepDOWN(p, nv, float(t[p]), float(t[nv]), float(z_mm[p]-z_mm[nv]), dur,
                                 v_min, v_mean, t_to_vmin, t_neg_acc, p_min, p_mean, work))
        # SEATED: desde nv hasta justo antes del siguiente levantarse (primer cruce vel neg->pos tras nv)
        if i+1 < len(peaks_sorted):
            idx_start = nv
            # buscar primer cruce neg->pos de vel tras nv
            sign = np.sign(vel_mm_s)
            sign[np.isnan(sign)] = 0
            zc = np.where(np.diff(sign[nv:]) > 0)[0]
            if len(zc):
                idx_end = nv + int(zc[0])
            else:
                idx_end = idx_start
            idx_end = min(idx_end, peaks_sorted[i+1])
            seg_v2 = vel_mm_s[idx_start:idx_end+1]
            seg_a2 = acc_m_s2[idx_start:idx_end+1] * 1000.0
            phases_seated.append(PhaseSEATED(idx_start, idx_end, float(t[idx_start]), float(t[idx_end]),
                                             float((idx_end-idx_start)*dt),
                                             float(np.nanmean(np.abs(seg_v2))) if seg_v2.size else np.nan,
                                             float(np.nanmean(np.abs(seg_a2))) if seg_a2.size else np.nan))

    # ---------------- Salidas ----------------
    out_prefix = out_prefix or (file_path.rsplit('.xlsx', 1)[0] + '_analysis3p')
    out_excel = out_prefix + '.xlsx'
    out_plot = out_prefix + '_segmentation.png'
    out_params = out_prefix + '_params.json'

    # Escribir en el Excel de origen + hojas nuevas
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        start_col_idx = column_index_from_string(start_col_excel)
        start_row = 1
        out_cols = pd.DataFrame({
            'BCM_z_mm (orig)': z_mm,
            'BCM_z_vel_c7 (mm/s)': vel_mm_s,
            'BCM_z_acc_c7 (mm/s^2)': acc_m_s2 * 1000.0,
            'Power_mg_v (W)': P_w,
        })
        for j, col in enumerate(out_cols.columns, start=start_col_idx):
            ws.cell(row=start_row, column=j, value=col)
        for i in range(len(z_mm)):
            for j, col in enumerate(out_cols.columns, start=start_col_idx):
                val = out_cols.iloc[i, j-start_col_idx]
                if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                    val = None
                ws.cell(row=start_row+1+i, column=j, value=val)

        # UP
        if 'STS Events (UP)' in wb.sheetnames:
            del wb['STS Events (UP)']
        ws_up = wb.create_sheet('STS Events (UP)')
        cols_up = ['rep','idx_trough','idx_peak','t_trough_s','t_peak_s','amp_up_mm','ok_up',
                   'dur_up_s','vmax_up_mm_s','vmean_up_mm_s','t_to_vmax_up_s','t_pos_acc_up_s',
                   'pmax_up_W','pmean_up_W','work_up_J']
        for j, c in enumerate(cols_up, start=1):
            ws_up.cell(row=1, column=j, value=c)
        for i, r in enumerate(reps_up, start=2):
            row = [i-1, r.idx_trough, r.idx_peak, r.t_trough, r.t_peak, r.amp_mm, r.ok, r.dur_s,
                   r.v_max_mms, r.v_mean_mms, r.t_to_vmax_s, r.t_pos_acc_s, r.p_max_w, r.p_mean_w, r.work_j]
            for j, val in enumerate(row, start=1):
                ws_up.cell(row=i, column=j, value=val)

        # DOWN
        if 'STS Events (DOWN)' in wb.sheetnames:
            del wb['STS Events (DOWN)']
        ws_dn = wb.create_sheet('STS Events (DOWN)')
        cols_dn = ['rep','idx_peak','idx_trough_next','t_peak_s','t_trough_next_s','amp_down_mm','dur_down_s',
                   'vmin_down_mm_s','vmean_down_mm_s','t_to_vmin_down_s','t_neg_acc_down_s','pmin_down_W','pmean_down_W','work_down_J']
        for j, c in enumerate(cols_dn, start=1):
            ws_dn.cell(row=1, column=j, value=c)
        for i, r in enumerate(reps_down, start=2):
            row = [i-1, r.idx_peak, r.idx_trough_next, r.t_peak, r.t_trough_next, r.amp_mm, r.dur_s,
                   r.v_min_mms, r.v_mean_mms, r.t_to_vmin_s, r.t_neg_acc_s, r.p_min_w, r.p_mean_w, r.work_j]
            for j, val in enumerate(row, start=1):
                ws_dn.cell(row=i, column=j, value=val)

        # SEATED
        if 'STS Phases (SEATED)' in wb.sheetnames:
            del wb['STS Phases (SEATED)']
        ws_st = wb.create_sheet('STS Phases (SEATED)')
        cols_st = ['phase','idx_start','idx_end','t_start_s','t_end_s','dur_s','v_abs_mean_mm_s','a_abs_mean_mm_s2']
        for j, c in enumerate(cols_st, start=1):
            ws_st.cell(row=1, column=j, value=c)
        for i, r in enumerate(phases_seated, start=2):
            row = ['SEATED', r.idx_start, r.idx_end, r.t_start, r.t_end, r.dur_s, r.v_abs_mean_mms, r.a_abs_mean_mms2]
            for j, val in enumerate(row, start=1):
                ws_st.cell(row=i, column=j, value=val)

        # Summary
        if 'STS Summary' in wb.sheetnames:
            del wb['STS Summary']
        ws_sm = wb.create_sheet('STS Summary')
        n_up = len(reps_up)
        n_ok_up = int(sum(r.ok for r in reps_up))
        amp_up = np.array([r.amp_mm for r in reps_up], float)
        dur_up = np.array([r.dur_s for r in reps_up], float)
        vmax_up = np.array([r.v_max_mms for r in reps_up], float)
        vmean_up = np.array([r.v_mean_mms for r in reps_up], float)
        t_vmax_up = np.array([r.t_to_vmax_s for r in reps_up], float)
        t_posacc_up = np.array([r.t_pos_acc_s for r in reps_up], float)
        work_up = np.array([r.work_j for r in reps_up if r.work_j is not None], float) if any(r.work_j is not None for r in reps_up) else None
        pmax_up = np.array([r.p_max_w for r in reps_up if r.p_max_w is not None], float) if any(r.p_max_w is not None for r in reps_up) else None
        pmean_up = np.array([r.p_mean_w for r in reps_up if r.p_mean_w is not None], float) if any(r.p_mean_w is not None for r in reps_up) else None

        dur_dn = np.array([r.dur_s for r in reps_down], float) if reps_down else None
        vmin_dn = np.array([r.v_min_mms for r in reps_down], float) if reps_down else None
        vmean_dn = np.array([r.v_mean_mms for r in reps_down], float) if reps_down else None
        t_vmin_dn = np.array([r.t_to_vmin_s for r in reps_down], float) if reps_down else None
        t_negacc_dn = np.array([r.t_neg_acc_s for r in reps_down], float) if reps_down else None
        work_dn = np.array([r.work_j for r in reps_down if r.work_j is not None], float) if any((r.work_j is not None) for r in reps_down) else None
        pmin_dn = np.array([r.p_min_w for r in reps_down if r.p_min_w is not None], float) if any((r.p_min_w is not None) for r in reps_down) else None
        pmean_dn = np.array([r.p_mean_w for r in reps_down if r.p_mean_w is not None], float) if any((r.p_mean_w is not None) for r in reps_down) else None

        dur_se = np.array([r.dur_s for r in phases_seated], float) if phases_seated else None
        vabs_se = np.array([r.v_abs_mean_mms for r in phases_seated], float) if phases_seated else None
        aabs_se = np.array([r.a_abs_mean_mms2 for r in phases_seated], float) if phases_seated else None

        summary = {
            'n_reps_up_total': n_up,
            'n_reps_up_ok_threshold': n_ok_up,
            'ok_threshold_frac': ok_th,
            'amp_up_max_mm': float(np.nanmax(amp_up)) if n_up else None,
            'amp_up_mean_mm': float(np.nanmean(amp_up)) if n_up else None,
            'amp_up_sd_mm': float(np.nanstd(amp_up, ddof=1)) if n_up > 1 else None,
            'dur_up_mean_s': float(np.nanmean(dur_up)) if n_up else None,
            'vmax_up_mean_mm_s': float(np.nanmean(vmax_up)) if n_up else None,
            'vmean_up_mean_mm_s': float(np.nanmean(vmean_up)) if n_up else None,
            't_to_vmax_up_mean_s': float(np.nanmean(t_vmax_up)) if n_up else None,
            't_pos_acc_up_mean_s': float(np.nanmean(t_posacc_up)) if n_up else None,
            'work_up_sum_J': float(np.nansum(work_up)) if work_up is not None else None,
            'pmax_up_mean_W': float(np.nanmean(pmax_up)) if pmax_up is not None else None,
            'pmean_up_mean_W': float(np.nanmean(pmean_up)) if pmean_up is not None else None,
            'dur_down_mean_s': float(np.nanmean(dur_dn)) if dur_dn is not None else None,
            'vmin_down_mean_mm_s': float(np.nanmean(vmin_dn)) if vmin_dn is not None else None,
            'vmean_down_mean_mm_s': float(np.nanmean(vmean_dn)) if vmean_dn is not None else None,
            't_to_vmin_down_mean_s': float(np.nanmean(t_vmin_dn)) if t_vmin_dn is not None else None,
            't_neg_acc_down_mean_s': float(np.nanmean(t_negacc_dn)) if t_negacc_dn is not None else None,
            'work_down_sum_J': float(np.nansum(work_dn)) if work_dn is not None else None,
            'pmin_down_mean_W': float(np.nanmean(pmin_dn)) if pmin_dn is not None else None,
            'pmean_down_mean_W': float(np.nanmean(pmean_dn)) if pmean_dn is not None else None,
            'dur_seated_mean_s': float(np.nanmean(dur_se)) if dur_se is not None else None,
            'vabs_seated_mean_mm_s': float(np.nanmean(vabs_se)) if vabs_se is not None else None,
            'aabs_seated_mean_mm_s2': float(np.nanmean(aabs_se)) if aabs_se is not None else None,
            'bcm_z_col': bcm_z_col,
            'mass_kg': mass_kg,
            'params_min_sep_s': used_min_sep,
            'params_prom_frac': used_prom,
            'power_model': p_model,
        }
        r = 1
        for k, v in summary.items():
            ws_sm.cell(row=r, column=1, value=k)
            ws_sm.cell(row=r, column=2, value=v)
            r += 1

        wb.save(out_excel)
    except Exception as e:
        # Fallback: Excel nuevo con todo
        with pd.ExcelWriter(out_excel, engine='openpyxl') as xw:
            df_out = df.copy()
            df_out['BCM_z_mm (orig)'] = z_mm
            df_out['BCM_z_vel_c7 (mm/s)'] = vel_mm_s
            df_out['BCM_z_acc_c7 (mm/s^2)'] = acc_m_s2 * 1000.0
            df_out['Power_mg_v (W)'] = P_w
            df_out.to_excel(xw, sheet_name=sheet_name, index=False)
            # UP
            df_up = pd.DataFrame([{
                'rep': i+1,
                'idx_trough': r.idx_trough,
                'idx_peak': r.idx_peak,
                't_trough_s': r.t_trough,
                't_peak_s': r.t_peak,
                'amp_up_mm': r.amp_mm,
                'ok_up': r.ok,
                'dur_up_s': r.dur_s,
                'vmax_up_mm_s': r.v_max_mms,
                'vmean_up_mm_s': r.v_mean_mms,
                't_to_vmax_up_s': r.t_to_vmax_s,
                't_pos_acc_up_s': r.t_pos_acc_s,
                'pmax_up_W': r.p_max_w,
                'pmean_up_W': r.p_mean_w,
                'work_up_J': r.work_j,
            } for i, r in enumerate(reps_up)])
            df_up.to_excel(xw, sheet_name='STS Events (UP)', index=False)
            # DOWN
            df_dn = pd.DataFrame([{
                'rep': i+1,
                'idx_peak': r.idx_peak,
                'idx_trough_next': r.idx_trough_next,
                't_peak_s': r.t_peak,
                't_trough_next_s': r.t_trough_next,
                'amp_down_mm': r.amp_mm,
                'dur_down_s': r.dur_s,
                'vmin_down_mm_s': r.v_min_mms,
                'vmean_down_mm_s': r.v_mean_mms,
                't_to_vmin_down_s': r.t_to_vmin_s,
                't_neg_acc_down_s': r.t_neg_acc_s,
                'pmin_down_W': r.p_min_w,
                'pmean_down_W': r.p_mean_w,
                'work_down_J': r.work_j,
            } for i, r in enumerate(reps_down)])
            df_dn.to_excel(xw, sheet_name='STS Events (DOWN)', index=False)
            # SEATED
            df_se = pd.DataFrame([{
                'phase': 'SEATED',
                'idx_start': r.idx_start,
                'idx_end': r.idx_end,
                't_start_s': r.t_start,
                't_end_s': r.t_end,
                'dur_s': r.dur_s,
                'v_abs_mean_mm_s': r.v_abs_mean_mms,
                'a_abs_mean_mm_s2': r.a_abs_mean_mms2,
            } for r in phases_seated])
            df_se.to_excel(xw, sheet_name='STS Phases (SEATED)', index=False)
            # Summary con error
            pd.DataFrame({'key':['error'], 'value':[str(e)]}).to_excel(xw, sheet_name='STS Summary', index=False)

    # Gráfico global
    if make_plot:
        fig, ax1 = plt.subplots(figsize=(12, 6))
        ax2 = ax1.twinx()
        ax1.plot(t, z_mm, color='tab:blue', label='BCM Z (mm)')
        ax2.plot(t, vel_mm_s, color='tab:orange', alpha=0.7, label='Vel Z (mm/s)')
        for k, ru in enumerate(reps_up, start=1):
            ax1.axvline(ru.t_trough, color='royalblue', linestyle='--', alpha=0.6)
            ax1.axvline(ru.t_peak, color='crimson', linestyle='--', alpha=0.6)
            ax1.text(ru.t_peak, z_mm[ru.idx_peak], f'{k}', color='crimson', fontsize=8, ha='left', va='bottom')
            if ru.ok == 1:
                ax1.fill_between([ru.t_trough, ru.t_peak], [z_mm[ru.idx_trough], z_mm[ru.idx_trough]], [z_mm[ru.idx_peak], z_mm[ru.idx_peak]], color='green', alpha=0.08)
        # Sentarse (rojo suave)
        for rd in reps_down:
            ax1.fill_between([rd.t_peak, rd.t_trough_next], [z_mm[rd.idx_peak], z_mm[rd.idx_peak]], [z_mm[rd.idx_trough_next], z_mm[rd.idx_trough_next]], color='red', alpha=0.05)
        # Sentado (gris)
        for rs in phases_seated:
            ax1.axvspan(rs.t_start, rs.t_end, color='gray', alpha=0.08)
        ax1.set_xlabel('Tiempo (s)')
        ax1.set_ylabel('Desplazamiento (mm)')
        ax2.set_ylabel('Velocidad (mm/s)')
        ax1.set_title('Segmentación STS: Levantarse / Sentarse / Sentado (BCM Z)')
        l1,_ = ax1.get_legend_handles_labels()
        l2,_ = ax2.get_legend_handles_labels()
        ax1.legend(l1+l2, [h.get_label() for h in l1+l2], loc='upper left')
        plt.tight_layout()
        fig.savefig(out_plot, dpi=150)
        plt.close(fig)

    # Gráficos por repetición
    per_rep_dir = None
    if per_rep_plots and len(reps_up):
        per_rep_dir = out_prefix + '_perrep'
        os.makedirs(per_rep_dir, exist_ok=True)
        for i, ru in enumerate(reps_up):
            rd = reps_down[i] if i < len(reps_down) else None
            rs = phases_seated[i] if i < len(phases_seated) else None
            t0 = ru.idx_trough
            t1 = ru.idx_peak
            t2 = rd.idx_trough_next if rd else min(ru.idx_peak + int(0.5/dt), len(t)-1)
            lo = max(0, t0 - int(0.5/dt))
            hi = min(len(t)-1, (rs.idx_end if rs else t2) + int(0.5/dt))
            fig, ax = plt.subplots(2, 1, figsize=(10,6), sharex=True)
            ax[0].plot(t[lo:hi+1], z_mm[lo:hi+1], 'b-', label='BCM Z (mm)')
            ax[1].plot(t[lo:hi+1], vel_mm_s[lo:hi+1], 'orange', label='Vel Z (mm/s)')
            ax[0].axvspan(t[t0], t[t1], color='green', alpha=0.1, label='Levantarse')
            if rd:
                ax[0].axvspan(t[t1], t[t2], color='red', alpha=0.08, label='Sentarse')
            if rs:
                ax[0].axvspan(t[rs.idx_start], t[rs.idx_end], color='gray', alpha=0.08, label='Sentado')
            ax[0].legend(loc='best')
            ax[1].set_xlabel('Tiempo (s)')
            ax[0].set_ylabel('z (mm)')
            ax[1].set_ylabel('vel (mm/s)')
            ax[0].set_title(f'Rep {i+1}: Levantarse / Sentarse / Sentado')
            plt.tight_layout()
            fig.savefig(os.path.join(per_rep_dir, f'rep_{i+1:02d}.png'), dpi=130)
            plt.close(fig)

    # Parámetros (JSON)
    params = {
        'file_path': file_path,
        'sheet_name': sheet_name,
        'time_col': time_col,
        'bcm_z_col': bcm_z_col,
        'dt': dt,
        'min_sep_s': used_min_sep,
        'prom_frac': used_prom,
        'ok_threshold': ok_th,
        'mass_kg': mass_kg,
        'power_model': p_model,
    }
    with open(out_params, 'w', encoding='utf-8') as f:
        json.dump(params, f, ensure_ascii=False, indent=2)

    return out_excel, out_plot, out_params


def main():
    parser = argparse.ArgumentParser(description='Herramienta de análisis STS con 3 fases (BCM Z)')
    parser.add_argument('--input', required=True, help='Ruta del archivo .xlsx de entrada')
    parser.add_argument('--sheet', default='Reducido', help='Hoja a procesar (por defecto: Reducido)')
    parser.add_argument('--time-col', default='time', help='Nombre de la columna de tiempo (por defecto: time)')
    parser.add_argument('--bcm-z-col', default=None, help='Nombre exacto de BCM Z (mm); si no, se autodetecta')
    parser.add_argument('--start-col', default='CW', help='Columna inicial para escribir (por defecto: CW)')
    parser.add_argument('--mass-kg', type=float, default=None, help='Masa corporal (kg) para P=m*g*v y trabajo')
    parser.add_argument('--min-sep', type=float, default=None, help='Separación mínima entre reps (s). Si no, auto-tune')
    parser.add_argument('--prom-frac', type=float, default=None, help='Prominencia mínima relativa (0-1). Si no, auto-tune')
    parser.add_argument('--ok-th', type=float, default=0.85, help='Umbral relativo para marcar OK (por defecto: 0.85)')
    parser.add_argument('--no-plot', action='store_true', help='No generar gráfico general')
    parser.add_argument('--no-per-rep', action='store_true', help='No generar gráficos por repetición')
    parser.add_argument('--out-prefix', default=None, help='Prefijo de salida (sin extensión)')

    args = parser.parse_args()

    excel, plot, params = run_tool(
        file_path=args.input,
        out_prefix=args.out_prefix,
        sheet_name=args.sheet,
        time_col=args.time_col,
        bcm_z_col=args.bcm_z_col,
        start_col_excel=args.start_col,
        mass_kg=args.mass_kg,
        min_sep_s=args.min_sep,
        prom_frac=args.prom_frac,
        ok_th=args.ok_th,
        make_plot=(not args.no_plot),
        per_rep_plots=(not args.no_per_rep),
    )

    print('\nLISTO')
    print('Excel:', excel)
    print('Gráfico:', plot)
    print('Parámetros:', params)


if __name__ == '__main__':
    main()
