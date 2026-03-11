# sts_analysis_tool_v3.py
# -*- coding: utf-8 -*-
"""
Versión v3:
- Detección exacta estilo Excel (ventana positiva/negativa de 30 y umbral vel ±0.1 m/s).
- Marcadores 'Acc Phases' y tiempos de fase por XLOOKUP equivalente:
  Concéntrica = [(r-1)*3+1 → +1], Excéntrica = [(r-1)*3+2 → +1], Sentado = [(r-1)*3+3 → +1 o fin].
- Salida en 3 hojas: Hoja1_Variables, Hoja2_Parametros_Grafica (con imagen) y
  Hoja3_Kinematic_&_Forces_like (3 filas por repetición).
- Métricas por fase: rango de BCM-Z (m), estadísticas de vel BCM-Z (m/s),
  potencia/trabajo (P≈m*g*v) si se da --mass-kg.
"""

import argparse
import os
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ---------- utilidades ----------
def centered_slope(y: np.ndarray, dt: float, half_window: int = 3) -> np.ndarray:
    y = np.asarray(y, dtype=float)
    dy = np.full_like(y, np.nan)
    w = half_window
    for i in range(w, len(y) - w):
        dy[i] = (y[i + w] - y[i - w]) / (2 * w * dt)
    return dy

def trapz_manual(y: np.ndarray, dx: float) -> float:
    y = np.asarray(y, dtype=float)
    if y.size < 2:
        return 0.0
    s = 0.0
    for i in range(1, y.size):
        a = y[i-1]; b = y[i]
        if np.isfinite(a) and np.isfinite(b):
            s += 0.5*(a+b)*dx
    return float(s)

@dataclass
class RepEvents:
    rep: int
    start_idx: int
    peak_idx: Optional[int]
    ecc_end_idx: Optional[int]

class FeatureComputer:
    def __init__(self, df: pd.DataFrame, time_col: str):
        self.df = df
        self.time = df[time_col].astype(float).to_numpy()
        self.cols = df.columns

    def _phase_slice(self, t0: float, t1: float) -> slice:
        i0 = int(np.argmin(np.abs(self.time - t0)))
        i1 = int(np.argmin(np.abs(self.time - t1)))
        if i1 < i0: i0, i1 = i1, i0
        return slice(i0, i1+1)

    def disp_range_m(self, col: str, t0: float, t1: float):
        if col not in self.cols: return (None,None,None)
        sl = self._phase_slice(t0, t1)
        vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        if not np.any(np.isfinite(vals)): return (None,None,None)
        mx = float(np.nanmax(vals))/1000.0
        mn = float(np.nanmin(vals))/1000.0
        return (mx, mn, mx-mn)

    def vel_stats(self, col: Optional[str], arr_fallback: Optional[np.ndarray], t0: float, t1: float):
        sl = self._phase_slice(t0, t1)
        if col and (col in self.cols):
            vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        elif arr_fallback is not None:
            vals = arr_fallback[sl]
        else:
            return (None,None,None)
        if not np.any(np.isfinite(vals)): return (None,None,None)
        return (float(np.nanmean(vals)), float(np.nanmax(vals)), float(np.nanmin(vals)))

    def power_work_stats(self, power_col: Optional[str], vel_col: Optional[str],
                         mass_kg: Optional[float], t0: float, t1: float):
        sl = self._phase_slice(t0, t1)
        if power_col and (power_col in self.cols):
            p = pd.to_numeric(self.df[power_col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        elif mass_kg is not None and vel_col and (vel_col in self.cols):
            g = 9.80665
            v = pd.to_numeric(self.df[vel_col].iloc[sl], errors='coerce').to_numpy(dtype=float)
            p = mass_kg * g * v
        else:
            return (None,None,None)
        if not np.any(np.isfinite(p)): return (None,None,None)
        dt = float(pd.Series(np.diff(self.time)).median())
        return (float(np.nanmean(p)), float(np.nanmax(p)), trapz_manual(p[np.isfinite(p)], dt))

# ---------- núcleo v3 ----------
def run_v3(
    file_path: str,
    sheet_name: str = 'Reducido',
    time_col: str = 'time',
    mass_kg: Optional[float] = None,
    window: int = 30,
    n_positive: int = 30,
    vel_th_m_s: float = 0.1,
    half_window_derivative: int = 3,
    out_prefix: Optional[str] = None,
    close_last_seated_at_end: bool = True,
):
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    t = df[time_col].astype(float).to_numpy()
    dt = float(pd.Series(np.diff(t)).median())

    # detectar BCM-Z (mm)
    disp_cols = [c for c in df.columns if 'body center of mass-z' in str(c).lower()]
    if not disp_cols: raise ValueError('No se encontró columna BCM Z (mm)')
    disp_col = disp_cols[0]
    z_mm = pd.to_numeric(df[disp_col], errors='coerce').to_numpy(dtype=float)
    z_m = z_mm/1000.0

    vel_m_s = centered_slope(z_m, dt, half_window_derivative)
    acc_m_s2 = centered_slope(vel_m_s, dt, half_window_derivative)
    power_W = mass_kg*9.80665*vel_m_s if mass_kg is not None else np.full_like(vel_m_s, np.nan)

    # vector de desplazamiento; ventanas; umbrales
    n = len(z_mm)
    dz = np.diff(z_mm, prepend=z_mm[0])
    vec = np.zeros(n, float); vec[dz>0]=1; vec[dz<0]=-1

    fut = np.full(n, np.nan); prv = np.full(n, np.nan)
    for i in range(n):
        if i+window <= n: fut[i]=np.nansum(vec[i:i+window])
        if i-window+1 >= 0: prv[i]=np.nansum(vec[i-window+1:i+1])

    vup  = (vel_m_s >  vel_th_m_s)
    vdn  = (vel_m_s < -vel_th_m_s)

    start = np.zeros(n, int)
    for i in range(n):
        if np.isnan(fut[i]): continue
        cond = (fut[i]>=n_positive) and vup[i]
        prev = (fut[i-1]>=n_positive and vup[i-1]) if (i>0 and not np.isnan(fut[i-1])) else False
        if cond and not prev: start[i]=10000

    peak = np.zeros(n, int)
    for i in range(n):
        if i-window<0 or i+window>=n: continue
        if (z_mm[i] > np.nanmax(z_mm[i-window:i])) and (z_mm[i] > np.nanmax(z_mm[i+1:i+1+window])):
            peak[i]=10000

    end  = np.zeros(n, int)
    for i in range(n):
        if np.isnan(prv[i]): continue
        cond = (prv[i]<=-n_positive) and vdn[i]
        nxt = (prv[i+1]<=-n_positive and vdn[i+1]) if (i+1<n and not np.isnan(prv[i+1])) else False
        if cond and not nxt: end[i]=10000

    phases_event = ((start==10000)|(peak==10000)|(end==10000)).astype(int)
    # marcadores acumulados exactos al estilo HI=SUM($HH$2:HHi)
    acc = np.zeros(n, int)
    curr=0
    s_idx = list(np.where(start==10000)[0])
    p_idx = list(np.where(peak==10000)[0])
    e_idx = list(np.where(end==10000)[0])
    pi=0; ei=0
    rep_events: List[RepEvents] = []
    for r, s in enumerate(s_idx, start=1):
        while pi < len(p_idx) and p_idx[pi] <= s: pi+=1
        p = p_idx[pi] if pi < len(p_idx) else None
        if p is not None: pi+=1
        e = None
        if p is not None:
            while ei < len(e_idx) and e_idx[ei] <= p: ei+=1
            e = e_idx[ei] if ei < len(e_idx) else None
            if e is not None: ei+=1
        rep_events.append(RepEvents(r, s, p, e))
        # asignar marcadores en orden
        curr+=1; acc[s]=curr
        if p is not None: curr+=1; acc[p]=curr
        if e is not None: curr+=1; acc[e]=curr
    for i in range(1,n):
        if acc[i]==0: acc[i]=acc[i-1]

    marker_to_time = {}
    for m in range(1, int(np.max(acc))+1):
        idx = int(np.argmax(acc==m)) if np.any(acc==m) else None
        if idx is not None and acc[idx]==m:
            marker_to_time[m] = float(t[idx])

    # Hoja1
    hoja1 = df.copy()
    hoja1['Disp_BCM_Z_mm']=z_mm
    hoja1['Vel_BCM_Z_m_s']=vel_m_s
    hoja1['Acc_BCM_Z_m_s2']=acc_m_s2
    hoja1['Power_BCM_Z_W']=power_W
    hoja1['Conc Start']=np.where(start==10000,10000,0)
    hoja1['Conc-Exc']=np.where(peak==10000,10000,0)
    hoja1['Ecc End']=np.where(end==10000,10000,0)
    hoja1['Phases']=phases_event
    hoja1['Acc Phases']=acc

    # Hoja3 tipo Kinematic_&_Forces
    try:
        meta_df = pd.read_excel(file_path, sheet_name='MetaData_&_Parameters', engine='openpyxl', header=None)
        meta = {
            'Codigo': meta_df.iloc[1,1] if meta_df.shape[0]>1 and meta_df.shape[1]>1 else None,
            'Altura': meta_df.iloc[2,1] if meta_df.shape[0]>2 and meta_df.shape[1]>1 else None,
            'Peso_kg': meta_df.iloc[3,1] if meta_df.shape[0]>3 and meta_df.shape[1]>1 else None,
            'Altura_silla': meta_df.iloc[6,1] if meta_df.shape[0]>6 and meta_df.shape[1]>1 else None,
            'Test': meta_df.iloc[4,1] if meta_df.shape[0]>4 and meta_df.shape[1]>1 else None,
            'Fecha_Test': meta_df.iloc[5,1] if meta_df.shape[0]>5 and meta_df.shape[1]>1 else None,
        }
    except Exception:
        meta = {'Codigo':None,'Altura':None,'Peso_kg':mass_kg,'Altura_silla':None,'Test':None,'Fecha_Test':None}

    feature = FeatureComputer(df, time_col)
    bcmz_col = disp_col
    # intentar localizar un vel_BCM_Z nativo si existiera
    vel_bcmz_col = next((c for c in df.columns if 'vel_bcm_z' in str(c).lower() or 'veloc_bcm_z' in str(c).lower()), None)

    rows=[]
    def add_phase(rep_idx, fase_name, m_start, m_end_opt):
        if m_start not in marker_to_time: return
        t0 = marker_to_time[m_start]
        t1 = marker_to_time[m_end_opt] if (m_end_opt in marker_to_time) else (float(t[-1]) if close_last_seated_at_end else t0)
        tf = max(0.0, t1 - t0)
        row = {
            'Código': meta['Codigo'], 'Altura': meta['Altura'], 'Peso_kg': meta['Peso_kg'],
            'Altura de la silla': meta['Altura_silla'], 'Test': meta['Test'], 'Fecha Test': meta['Fecha_Test'],
            'Repeticion': rep_idx, 'Fase': fase_name, 'Inicio_Tiempo': t0, 'Final_Tiempo': t1, 'Tiempo fase': tf,
        }
        mx, mn, rg = feature.disp_range_m(bcmz_col, t0, t1)
        row['Max Disp_BCM_Z (m)']=mx; row['Min Disp_BCM_Z (m)']=mn; row['Range Disp_BCM_Z (m)']=rg
        mean_v, max_v, min_v = feature.vel_stats(vel_bcmz_col, vel_m_s, t0, t1)
        row['Mean Vel_BCM_Z (m/s)']=mean_v; row['Max Vel_BCM_Z (m/s)']=max_v; row['Min Vel_BCM_Z (m/s)']=min_v
        meanP, maxP, work = feature.power_work_stats(None, vel_bcmz_col, meta['Peso_kg'] if meta['Peso_kg'] is not None else mass_kg, t0, t1)
        row['Mean Power_BCM_Z (W)']=meanP; row['Max Power_BCM_Z (W)']=maxP; row['Mech Work_BCM_Z (J)']=work
        rows.append(row)

    for ev in rep_events:
        r=ev.rep
        m1=(r-1)*3+1; m2=(r-1)*3+2; m3=(r-1)*3+3
        add_phase(r, 'Concéntrica', m1, m2)
        add_phase(r, 'Excéntrica',  m2, m3)
        add_phase(r, 'Sentado',     m3, m3+1)

    hoja3 = pd.DataFrame(rows)

    # Hoja2 (parámetros + gráfica)
    out_prefix = out_prefix or (os.path.splitext(file_path)[0] + '_analysis_v3')
    plot_path = out_prefix + '_segmentation.png'
    try:
        fig, ax1 = plt.subplots(figsize=(12,6)); ax2 = ax1.twinx()
        ax1.plot(t, z_mm, 'b-', label='BCM Z (mm)')
        ax2.plot(t, vel_m_s, color='orange', alpha=0.7, label='Vel BCM Z (m/s)')
        colors={'Concéntrica':'green','Excéntrica':'red','Sentado':'gray'}
        for _, r in hoja3.iterrows():
            ax1.axvspan(r['Inicio_Tiempo'], r['Final_Tiempo'], color=colors.get(r['Fase'],'k'), alpha=0.06)
        sI=np.where(start==10000)[0]; pI=np.where(peak==10000)[0]; eI=np.where(end==10000)[0]
        if len(sI): ax1.scatter(t[sI], z_mm[sI], c='green', marker='^', label='Conc Start')
        if len(pI): ax1.scatter(t[pI], z_mm[pI], c='purple', marker='o', label='Conc-Exc')
        if len(eI): ax1.scatter(t[eI], z_mm[eI], c='red', marker='v', label='Ecc End')
        ax1.set_xlabel('Tiempo (s)'); ax1.set_ylabel('BCM Z (mm)'); ax2.set_ylabel('Vel BCM Z (m/s)')
        ax1.legend(loc='upper left'); plt.tight_layout(); fig.savefig(plot_path, dpi=150); plt.close(fig)
    except Exception:
        plot_path=None

    # Escribir Excel
    wb=Workbook(); ws1=wb.active; ws1.title='Hoja1_Variables'
    for j,col in enumerate(hoja1.columns, start=1):
        ws1.cell(1,j,str(col)).font=Font(bold=True)
    for i, (_, row) in enumerate(hoja1.iterrows(), start=2):
        for j,val in enumerate(row, start=1):
            if isinstance(val, float) and (np.isnan(val) or np.isinf(val)): val=None
            ws1.cell(i,j,val)
    for j in range(1, len(hoja1.columns)+1):
        ws1.column_dimensions[get_column_letter(j)].width= min(max(len(str(hoja1.columns[j-1]))+2,12),30)

    ws2=wb.create_sheet('Hoja2_Parametros_Grafica')
    params=[('Archivo entrada',file_path),('Hoja analizada',sheet_name),('Columna tiempo',time_col),
            ('Columna BCM Z',disp_col),('dt (s)',dt),('Ventana',window),
            ('N positivas/negativas',n_positive),('Umbral vel (m/s)',vel_th_m_s),
            ('Masa (kg)',mass_kg),('Deriv half-window',half_window_derivative),
            ('N starts',len(s_idx)),('N peaks',len(p_idx)),('N ecc_end',len(e_idx)),('N reps',len(rep_events))]
    for r,(k,v) in enumerate(params, start=1):
        ws2.cell(r,1,k).font=Font(bold=True); ws2.cell(r,2,v)
    if plot_path and os.path.exists(plot_path):
        img=XLImage(plot_path); img.width=1100; img.height=540; ws2.add_image(img,'D2')

    ws3=wb.create_sheet('Hoja3_Kinematic_&_Forces_like')
    if not hoja3.empty:
        for j,col in enumerate(hoja3.columns, start=1):
            ws3.cell(1,j,col).font=Font(bold=True)
        for i, (_, row) in enumerate(hoja3.iterrows(), start=2):
            for j, val in enumerate(row, start=1):
                if isinstance(val, float) and (np.isnan(val) or np.isinf(val)): val=None
                ws3.cell(i,j,val)
        for j in range(1, len(hoja3.columns)+1):
            ws3.column_dimensions[get_column_letter(j)].width = min(max(len(str(hoja3.columns[j-1]))+2,14),32)
    else:
        ws3['A1']='Sin repetición detectada'

    out_xlsx = out_prefix + '.xlsx'
    wb.save(out_xlsx)

    out_json = out_prefix + '_params.json'
    with open(out_json,'w',encoding='utf-8') as f:
        json.dump({'file_path':file_path,'sheet_name':sheet_name,'time_col':time_col,
                   'dt':dt,'window':window,'n_positive':n_positive,'vel_th_m_s':vel_th_m_s,
                   'mass_kg':mass_kg,'half_window_derivative':half_window_derivative}, f, ensure_ascii=False, indent=2)

    return out_xlsx, plot_path, out_json

if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='STS v3: fases por Acc Phases + Hoja3 por fase estilo Kinematic_&_Forces')
    ap.add_argument('--input', required=True)
    ap.add_argument('--sheet', default='Reducido')
    ap.add_argument('--time-col', default='time')
    ap.add_argument('--mass-kg', type=float, default=None)
    ap.add_argument('--window', type=int, default=30)
    ap.add_argument('--n-positive', type=int, default=30)
    ap.add_argument('--vel-th', type=float, default=0.1)
    ap.add_argument('--half-window-derivative', type=int, default=3)
    ap.add_argument('--out-prefix', default=None)
    args = ap.parse_args()

    xlsx, png, js = run_v3(
        file_path=args.input,
        sheet_name=args.sheet,
        time_col=args.time_col,
        mass_kg=args.mass_kg,
        window=args.window,
        n_positive=args.n_positive,
        vel_th_m_s=args.vel_th,
        half_window_derivative=args.half_window_derivative,
        out_prefix=args.out_prefix
    )
    print('\\nLISTO'); print('Excel:', xlsx); print('Gráfico:', png); print('Parámetros:', js)