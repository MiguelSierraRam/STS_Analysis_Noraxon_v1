# sts_analysis_tool_v3_3.py
# -*- coding: utf-8 -*-
"""
STS Analysis Tool v3.3 (BCM Z - Acc Phases - 3 Phases per rep)

This legacy script has been retrofitted with configuration and logging support
compatible with the newer `enhanced_v2` tool. You can provide a YAML config
file via `--config` and override parameters with CLI options. Outputs are
logged using the centralized logging infrastructure (`src/logger.py`).

Highlights vs v3.2:
- Batch mode: --batch data/input  (process all .xlsx inside; recursive off by default)
- Output directory: --out data/output  (created if missing)
- Skip already-processed files: if output Excel exists and is newer than input, it is skipped (unless --overwrite)
- CSV export of Hoja3 via --csv
- Keeps logic of v3.2 for Hoja3 metrics: BCM_Z, HIP_Z, EMG means, CoP_SD, Angles, Forces
- Integration via manual trapezoid (compatible NumPy 1.x / 2.x)

Single-file usage:
    python sts_analysis_tool_v3_3.py --input data/input/YourFile.xlsx --sheet Reducido --mass-kg 72
Batch usage:
    python sts_analysis_tool_v3_3.py --batch data/input --out data/output --mass-kg 72 --csv
"""

import argparse
import os
import json
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple

from src.config import load_config, get_config
from src.logger import LoggerManager, get_logger

logger = get_logger(__name__)

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ---------------- utilidades ----------------

def centered_slope(y: np.ndarray, dt: float, half_window: int = 3) -> np.ndarray:
    y = np.asarray(y, dtype=float)
    dy = np.full_like(y, np.nan)
    w = half_window
    for i in range(w, len(y) - w):
        dy[i] = (y[i + w] - y[i - w]) / (2 * w * dt)
    return dy

def trapz_manual(y: np.ndarray, dx: float) -> float:
    y = np.asarray(y, dtype=float)
    if y.size < 2:
        return 0.0
    s = 0.0
    for i in range(1, y.size):
        a = y[i-1]; b = y[i]
        if np.isfinite(a) and np.isfinite(b):
            s += 0.5 * (a + b) * dx
    return float(s)

@dataclass
class RepEvents:
    rep: int
    start_idx: int
    peak_idx: Optional[int]
    ecc_end_idx: Optional[int]

class PhaseComputer:
    def __init__(self, df: pd.DataFrame, time_col: str):
        self.df = df
        self.time = df[time_col].astype(float).to_numpy()
        self.cols = df.columns
        self.dt = float(pd.Series(np.diff(self.time)).median())

    def _slice(self, t0: float, t1: float) -> slice:
        i0 = int(np.argmin(np.abs(self.time - t0)))
        i1 = int(np.argmin(np.abs(self.time - t1)))
        if i1 < i0:
            i0, i1 = i1, i0
        return slice(i0, i1+1)

    def range_mm_to_m(self, col: str, t0: float, t1: float) -> Tuple[Optional[float], Optional[float], Optional[float]]:
        if col not in self.cols:
            return (None,None,None)
        sl = self._slice(t0, t1)
        vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        if not np.any(np.isfinite(vals)):
            return (None,None,None)
        mx = float(np.nanmax(vals))/1000.0
        mn = float(np.nanmin(vals))/1000.0
        return (mx, mn, mx-mn)

    def range_deg(self, col: str, t0: float, t1: float) -> Tuple[Optional[float], Optional[float], Optional[float]]:
        if col not in self.cols:
            return (None,None,None)
        sl = self._slice(t0, t1)
        vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        if not np.any(np.isfinite(vals)):
            return (None,None,None)
        mx = float(np.nanmax(vals))
        mn = float(np.nanmin(vals))
        return (mx, mn, mx-mn)

    def stats(self, col: Optional[str], arr_fallback: Optional[np.ndarray], t0: float, t1: float) -> Tuple[Optional[float], Optional[float], Optional[float]]:
        sl = self._slice(t0, t1)
        if col and (col in self.cols):
            vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        elif arr_fallback is not None:
            vals = arr_fallback[sl]
        else:
            return (None,None,None)
        if not np.any(np.isfinite(vals)):
            return (None,None,None)
        return (float(np.nanmean(vals)), float(np.nanmax(vals)), float(np.nanmin(vals)))

    def power_work(self, power_col: Optional[str], vel_col: Optional[str], mass_kg: Optional[float], t0: float, t1: float) -> Tuple[Optional[float], Optional[float], Optional[float]]:
        sl = self._slice(t0, t1)
        if power_col and (power_col in self.cols):
            p = pd.to_numeric(self.df[power_col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        elif mass_kg is not None and vel_col and (vel_col in self.cols):
            # aproximación P = m*g*v_z (válida para BCM z)
            g = 9.80665
            v = pd.to_numeric(self.df[vel_col].iloc[sl], errors='coerce').to_numpy(dtype=float)
            p = mass_kg * g * v
        else:
            return (None,None,None)
        if not np.any(np.isfinite(p)):
            return (None,None,None)
        mean_p = float(np.nanmean(p))
        max_p = float(np.nanmax(p))
        work = trapz_manual(p[np.isfinite(p)], self.dt)
        return (mean_p, max_p, work)

    def mean(self, col: str, t0: float, t1: float) -> Optional[float]:
        if col not in self.cols:
            return None
        sl = self._slice(t0, t1)
        vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        if not np.any(np.isfinite(vals)):
            return None
        return float(np.nanmean(vals))

    def stdev(self, col: str, t0: float, t1: float) -> Optional[float]:
        if col not in self.cols:
            return None
        sl = self._slice(t0, t1)
        vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        vals = vals[np.isfinite(vals)]
        if vals.size < 2:
            return None
        return float(np.nanstd(vals, ddof=1))

    def time_to_max(self, col: str, t0: float, t1: float) -> Tuple[Optional[float], Optional[float]]:
        if col not in self.cols:
            return (None, None)
        sl = self._slice(t0, t1)
        vals = pd.to_numeric(self.df[col].iloc[sl], errors='coerce').to_numpy(dtype=float)
        idx = np.nanargmax(vals) if np.any(np.isfinite(vals)) else None
        if idx is None:
            return (None, None)
        tt = float((idx) * self.dt)
        dur = float((sl.stop - sl.start - 1) * self.dt)
        pct = float(tt/dur*100.0) if dur > 0 else None
        return (tt, pct)

# ---------------- núcleo principal (por archivo) ----------------

def analyze_file(
    input_path: str,
    sheet_name: str,
    time_col: str,
    mass_kg: Optional[float],
    window: int,
    n_positive: int,
    vel_th_m_s: float,
    half_window_derivative: int,
    out_dir: str,
    csv_export: bool,
    close_last_seated_at_end: bool = True,
) -> Tuple[str, str, str, Optional[str]]:
    logger.info(f"Iniciando análisis: {input_path}")
    # salida y prefijos
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(input_path))[0]
    out_prefix = os.path.join(out_dir, base + '_analysis_v3_3')
    out_xlsx = out_prefix + '.xlsx'
    out_png  = out_prefix + '_segmentation.png'
    out_json = out_prefix + '_params.json'
    out_csv  = out_prefix + '_Hoja3.csv'

    # cargar
    df = pd.read_excel(input_path, sheet_name=sheet_name, engine='openpyxl')
    if time_col not in df.columns:
        raise ValueError('No se encontró la columna de tiempo: %s' % time_col)
    t = df[time_col].astype(float).to_numpy()
    dt = float(pd.Series(np.diff(t)).median())

    # BCM Z
    disp_bcmz = next((c for c in df.columns if 'body center of mass-z' in str(c).lower()), None)
    if disp_bcmz is None:
        raise ValueError('No se encontró la columna BCM Z (mm)')
    z_mm = pd.to_numeric(df[disp_bcmz], errors='coerce').to_numpy(dtype=float)
    z_m = z_mm/1000.0
    vel_bcm_m_s = centered_slope(z_m, dt, half_window_derivative)
    acc_bcm_m_s2 = centered_slope(vel_bcm_m_s, dt, half_window_derivative)

    # HIP Z
    disp_hipz = next((c for c in df.columns if 'hip rt-z' in str(c).lower()), None)
    vel_hipz_native = next((c for c in df.columns if 'veloc_hip_z' in str(c).lower() or 'vel hip z' in str(c).lower()), None)
    power_hipz_native = next((c for c in df.columns if 'mean power_hip_z' in str(c).lower()), None)
    work_hipz_native  = next((c for c in df.columns if 'mech work_hip_z' in str(c).lower()), None)

    # EMG confirmadas
    emg_cols = [
        'RT TIB.ANT. (%)','RT VLO (%)','RT RECTUS FEM. (%)',
        'RT MED. GASTRO (%)','RT SEMITEND. (%)','RT GLUT. MAX. (%)','RT LUMBAR ES (%)'
    ]
    emg_cols = [c for c in emg_cols if c in df.columns]

    # CoP SD confirmadas + crudas
    raw_cop_candidates = {
        'AP': ['CoP_Disp_X','CoP_AP','CoP_X'],
        'ML': ['CoP_Disp_Y','CoP_ML','CoP_Y'],
        'Result': ['CoP_Disp_R','CoP_Result','CoP_R']
    }

    # ÁNGULOS
    angle_candidates = [
        'Thoracic Flexion Fwd (deg)', 'Lumbar Flexion Fwd (deg)', 'Torso-Pelvic Flexion Fwd (deg)',
        'RT Hip Flexion (deg)', 'RT Knee Flexion (deg)', 'RT Ankle Dorsiflexion (deg)'
    ]
    for extra in df.columns:
        cl = str(extra).lower()
        if ('pitch' in cl or 'flexion' in cl) and ('deg' in cl) and (extra not in angle_candidates):
            angle_candidates.append(extra)
    angle_cols = [c for c in angle_candidates if c in df.columns]

    # FUERZAS
    force_sets = {
        'Vert': {'mean':'Vert F_Mean', 'max':'Vert F_Max'},
        'Horizon': {'mean':'Horizon F_Mean', 'max':'Horizon F_Max'},
        'Result': {'mean':'Result F_Mean', 'max':'Result F_Max'}
    }
    for k in list(force_sets.keys()):
        mcol = force_sets[k]['mean']; xcol = force_sets[k]['max']
        if (mcol not in df.columns) and (xcol not in df.columns):
            force_sets.pop(k)

    # Detección estilo Excel
    n = len(z_mm)
    dz = np.diff(z_mm, prepend=z_mm[0])
    vec = np.zeros(n, float); vec[dz>0]=1; vec[dz<0]=-1
    fut = np.full(n, np.nan); prv = np.full(n, np.nan)
    for i in range(n):
        if i+window <= n: fut[i]=np.nansum(vec[i:i+window])
        if i-window+1 >= 0: prv[i]=np.nansum(vec[i-window+1:i+1])

    # velocidad para umbrales (sobre z_m)
    vel_for_th = centered_slope(z_m, dt, half_window_derivative)
    vup = (vel_for_th >  vel_th_m_s)
    vdn = (vel_for_th < -vel_th_m_s)

    start = np.zeros(n, int)
    for i in range(n):
        if np.isnan(fut[i]):
            continue
        cond = (fut[i] >= n_positive) and vup[i]
        prev = (fut[i-1] >= n_positive and vup[i-1]) if (i>0 and not np.isnan(fut[i-1])) else False
        if cond and not prev:
            start[i] = 10000

    peak = np.zeros(n, int)
    for i in range(n):
        if i-window < 0 or i+window >= n:
            continue
        if (z_mm[i] > np.nanmax(z_mm[i-window:i])) and (z_mm[i] > np.nanmax(z_mm[i+1:i+1+window])):
            peak[i] = 10000

    end = np.zeros(n, int)
    for i in range(n):
        if np.isnan(prv[i]):
            continue
        cond = (prv[i] <= -n_positive) and vdn[i]
        nxt = (prv[i+1] <= -n_positive and vdn[i+1]) if (i+1<n and not np.isnan(prv[i+1])) else False
        if cond and not nxt:
            end[i] = 10000

    # marcadores Acc Phases + rep_events
    acc = np.zeros(n, int)
    curr=0
    s_idx = list(np.where(start==10000)[0])
    p_idx = list(np.where(peak==10000)[0])
    e_idx = list(np.where(end==10000)[0])
    pi=0; ei=0
    rep_events: List[RepEvents] = []
    for r, s in enumerate(s_idx, start=1):
        while pi < len(p_idx) and p_idx[pi] <= s: pi += 1
        p = p_idx[pi] if pi < len(p_idx) else None
        if p is not None: pi += 1
        e = None
        if p is not None:
            while ei < len(e_idx) and e_idx[ei] <= p: ei += 1
            e = e_idx[ei] if ei < len(e_idx) else None
            if e is not None: ei += 1
        rep_events.append(RepEvents(r, s, p, e))
        curr += 1; acc[s]=curr
        if p is not None: curr += 1; acc[p]=curr
        if e is not None: curr += 1; acc[e]=curr
    for i in range(1, n):
        if acc[i]==0: acc[i]=acc[i-1]

    marker_to_time: Dict[int, float] = {}
    for m in range(1, int(np.max(acc))+1):
        if np.any(acc==m):
            idx = int(np.argmax(acc==m))
            marker_to_time[m] = float(t[idx])

    # Hoja1
    power_bcm_W = mass_kg*9.80665*vel_bcm_m_s if mass_kg is not None else np.full_like(vel_bcm_m_s, np.nan)
    hoja1 = df.copy()
    hoja1['Disp_BCM_Z_mm'] = z_mm
    hoja1['Vel_BCM_Z_m_s'] = vel_bcm_m_s
    hoja1['Acc_BCM_Z_m_s2'] = acc_bcm_m_s2
    hoja1['Power_BCM_Z_W'] = power_bcm_W
    hoja1['Conc Start'] = np.where(start==10000,10000,0)
    hoja1['Conc-Exc']   = np.where(peak==10000,10000,0)
    hoja1['Ecc End']    = np.where(end==10000,10000,0)
    hoja1['Acc Phases'] = acc

    # Hoja3 (Kinematic_&_Forces_like)
    try:
        meta_df = pd.read_excel(input_path, sheet_name='MetaData_&_Parameters', engine='openpyxl', header=None)
        meta = {
            'Codigo': meta_df.iloc[1,1] if meta_df.shape[0]>1 and meta_df.shape[1]>1 else None,
            'Altura': meta_df.iloc[2,1] if meta_df.shape[0]>2 and meta_df.shape[1]>1 else None,
            'Peso_kg': meta_df.iloc[3,1] if meta_df.shape[0]>3 and meta_df.shape[1]>1 else mass_kg,
            'Altura_silla': meta_df.iloc[6,1] if meta_df.shape[0]>6 and meta_df.shape[1]>1 else None,
            'Test': meta_df.iloc[4,1] if meta_df.shape[0]>4 and meta_df.shape[1]>1 else None,
            'Fecha_Test': meta_df.iloc[5,1] if meta_df.shape[0]>5 and meta_df.shape[1]>1 else None,
        }
    except Exception:
        meta = {'Codigo':None,'Altura':None,'Peso_kg':mass_kg,'Altura_silla':None,'Test':None,'Fecha_Test':None}

    pc = PhaseComputer(df, time_col)

    def phase_times(rep_idx: int, phase: str):
        m1=(rep_idx-1)*3+1; m2=m1+1; m3=m1+2
        if phase=='Concéntrica':
            sm, em = m1, m2
        elif phase=='Excéntrica':
            sm, em = m2, m3
        else:  # Sentado
            sm, em = m3, m3+1
        if sm not in marker_to_time:
            return None
        t0 = marker_to_time[sm]
        t1 = marker_to_time[em] if em in marker_to_time else (float(t[-1]) if close_last_seated_at_end else t0)
        return t0, t1

    rows: List[Dict[str, object]] = []
    for ev in rep_events:
        for phase in ('Concéntrica','Excéntrica','Sentado'):
            ts = phase_times(ev.rep, phase)
            if ts is None: continue
            t0, t1 = ts
            base = {
                'Código': meta['Codigo'], 'Altura': meta['Altura'], 'Peso_kg': meta['Peso_kg'],
                'Altura de la silla': meta['Altura_silla'], 'Test': meta['Test'], 'Fecha Test': meta['Fecha_Test'],
                'Repeticion': ev.rep, 'Fase': phase, 'Inicio_Tiempo': t0, 'Final_Tiempo': t1,
                'Tiempo fase': max(0.0, t1-t0),
            }
            # BCM Z
            mx,mn,rg = pc.range_mm_to_m(disp_bcmz, t0, t1)
            base['BCM_Z_Max (m)']=mx; base['BCM_Z_Min (m)']=mn; base['BCM_Z_Range (m)']=rg
            mV, MV, mVn = pc.stats(None, vel_bcm_m_s, t0, t1)
            base['BCM_Z_Vel_Mean (m/s)']=mV; base['BCM_Z_Vel_Max (m/s)']=MV; base['BCM_Z_Vel_Min (m/s)']=mVn
            if meta['Peso_kg'] is not None:
                sl = pc._slice(t0, t1)
                P_series = meta['Peso_kg']*9.80665*vel_bcm_m_s[sl]
                if np.any(np.isfinite(P_series)):
                    base['BCM_Z_Power_Mean (W)'] = float(np.nanmean(P_series))
                    base['BCM_Z_Power_Max (W)']  = float(np.nanmax(P_series))
                    base['BCM_Z_Work (J)']       = trapz_manual(P_series[np.isfinite(P_series)], dt)

            # HIP Z
            if disp_hipz is not None:
                hmx,hmn,hrg = pc.range_mm_to_m(disp_hipz, t0, t1)
                base['HIP_Z_Max (m)']=hmx; base['HIP_Z_Min (m)']=hmn; base['HIP_Z_Range (m)']=hrg
                if vel_hipz_native and vel_hipz_native in df.columns:
                    hVm, hVx, hVn = pc.stats(vel_hipz_native, None, t0, t1)
                else:
                    hip_mm = pd.to_numeric(df[disp_hipz], errors='coerce').to_numpy(dtype=float)
                    hip_m  = hip_mm/1000.0
                    hip_vel = centered_slope(hip_m, dt, half_window_derivative)
                    hVm, hVx, hVn = pc.stats(None, hip_vel, t0, t1)
                base['HIP_Z_Vel_Mean (m/s)']=hVm; base['HIP_Z_Vel_Max (m/s)']=hVx; base['HIP_Z_Vel_Min (m/s)']=hVn
                if power_hipz_native and power_hipz_native in df.columns:
                    base['HIP_Z_Power_Mean (W)'] = pc.mean(power_hipz_native, t0, t1)
                    base['HIP_Z_Power_Max (W)']  = pc.stats(power_hipz_native, None, t0, t1)[1]
                if work_hipz_native and work_hipz_native in df.columns:
                    sl = pc._slice(t0, t1)
                    series = pd.to_numeric(df[work_hipz_native].iloc[sl], errors='coerce').to_numpy(dtype=float)
                    series = series[np.isfinite(series)]
                    base['HIP_Z_Work (J)'] = float(series[-1]-series[0]) if series.size>=2 else (float(series[0]) if series.size==1 else None)

            # EMG medios
            for emg in emg_cols:
                base[f'EMG Mean {emg}'] = pc.mean(emg, t0, t1)

            # CoP SD (preferimos SD ya calculada; si no, SD de crudas)
            if 'CoP_SD Displ_AP' in df.columns:
                base['CoP_SD_AP'] = pc.mean('CoP_SD Displ_AP', t0, t1)
            else:
                for cand in raw_cop_candidates['AP']:
                    if cand in df.columns:
                        base['CoP_SD_AP'] = pc.stdev(cand, t0, t1); break
            if 'CoP_SD Displ_ML' in df.columns:
                base['CoP_SD_ML'] = pc.mean('CoP_SD Displ_ML', t0, t1)
            else:
                for cand in raw_cop_candidates['ML']:
                    if cand in df.columns:
                        base['CoP_SD_ML'] = pc.stdev(cand, t0, t1); break
            if 'CoP_SD Displ_Result' in df.columns:
                base['CoP_SD_Result'] = pc.mean('CoP_SD Displ_Result', t0, t1)
            else:
                for cand in raw_cop_candidates['Result']:
                    if cand in df.columns:
                        base['CoP_SD_Result'] = pc.stdev(cand, t0, t1); break

            # ÁNGULOS
            for ang in angle_cols:
                aMx, aMn, aRg = pc.range_deg(ang, t0, t1)
                base[f'{ang} - Max (deg)'] = aMx
                base[f'{ang} - Min (deg)'] = aMn
                base[f'{ang} - Range (deg)'] = aRg

            # FUERZAS
            for key, mm in force_sets.items():
                mcol = mm.get('mean'); xcol = mm.get('max')
                if mcol in df.columns:
                    base[f'{key} Force Mean'] = pc.mean(mcol, t0, t1)
                if xcol in df.columns:
                    base[f'{key} Force Max'] = pc.mean(xcol, t0, t1)
                    tt, pct = pc.time_to_max(xcol, t0, t1)
                    base[f'Time to {key} F_Max (s)'] = tt
                    base[f'Time to {key} F_Max (%)'] = pct

            rows.append(base)

    hoja3 = pd.DataFrame(rows)

    # Hoja2 (parámetros + gráfica)
    plot_path = out_png
    try:
        fig, ax1 = plt.subplots(figsize=(12,6)); ax2=ax1.twinx()
        ax1.plot(t, z_mm, 'b-', label='BCM Z (mm)')
        ax2.plot(t, vel_bcm_m_s, color='orange', alpha=0.7, label='Vel BCM Z (m/s)')
        colors={'Concéntrica':'green','Excércica':'red','Sentado':'gray'}
        # sombreado por filas de Hoja3
        for _, r in hoja3.iterrows():
            ax1.axvspan(r['Inicio_Tiempo'], r['Final_Tiempo'], color=colors.get(r['Fase'],'k'), alpha=0.06)
        sI=np.where(start==10000)[0]; pI=np.where(peak==10000)[0]; eI=np.where(end==10000)[0]
        if len(sI): ax1.scatter(t[sI], z_mm[sI], c='green', marker='^', label='Conc Start')
        if len(pI): ax1.scatter(t[pI], z_mm[pI], c='purple', marker='o', label='Conc-Exc')
        if len(eI): ax1.scatter(t[eI], z_mm[eI], c='red', marker='v', label='Ecc End')
        ax1.set_xlabel('Tiempo (s)'); ax1.set_ylabel('BCM Z (mm)'); ax2.set_ylabel('Vel BCM Z (m/s)')
        ax1.legend(loc='upper left'); plt.tight_layout(); fig.savefig(plot_path, dpi=150); plt.close(fig)
    except Exception:
        plot_path=None

    # Escribir Excel
    wb=Workbook(); ws1=wb.active; ws1.title='Hoja1_Variables'
    for j,col in enumerate(hoja1.columns, start=1):
        ws1.cell(1,j,str(col)).font=Font(bold=True)
    for i, (_, row) in enumerate(hoja1.iterrows(), start=2):
        for j,val in enumerate(row, start=1):
            if isinstance(val, float) and (np.isnan(val) or np.isinf(val)): val=None
            ws1.cell(i,j,val)
    for j in range(1, len(hoja1.columns)+1):
        ws1.column_dimensions[get_column_letter(j)].width = min(max(len(str(hoja1.columns[j-1]))+2,12),30)

    ws2=wb.create_sheet('Hoja2_Parametros_Grafica')
    params=[('Archivo entrada',input_path),('Hoja analizada',sheet_name),('Columna tiempo',time_col),
            ('Columna BCM Z',disp_bcmz),('dt (s)',dt),('Ventana',window),
            ('N positivas/negativas',n_positive),('Umbral vel (m/s)',vel_th_m_s),
            ('Masa (kg)',mass_kg),('Deriv half-window',half_window_derivative),
            ('N starts',len(s_idx)),('N peaks',len(p_idx)),('N ecc_end',len(e_idx)),('N reps',len(rep_events))]
    for r,(k,v) in enumerate(params, start=1):
        ws2.cell(r,1,k).font=Font(bold=True); ws2.cell(r,2,v)
    if plot_path and os.path.exists(plot_path):
        img=XLImage(plot_path); img.width=1100; img.height=540; ws2.add_image(img,'D2')

    ws3=wb.create_sheet('Hoja3_Kinematic_&_Forces_like')
    if not hoja3.empty:
        for j,col in enumerate(hoja3.columns, start=1): ws3.cell(1,j,str(col)).font=Font(bold=True)
        for i, (_, row) in enumerate(hoja3.iterrows(), start=2):
            for j,val in enumerate(row, start=1):
                if isinstance(val, float) and (np.isnan(val) or np.isinf(val)): val=None
                ws3.cell(i,j,val)
        for j in range(1, len(hoja3.columns)+1):
            ws3.column_dimensions[get_column_letter(j)].width = min(max(len(str(hoja3.columns[j-1]))+2,14),36)
    else:
        ws3['A1']='Sin repetición detectada'

    wb.save(out_xlsx)

    # JSON y CSV
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump({'file_path':input_path,'sheet_name':sheet_name,'time_col':time_col,
                   'dt':dt,'window':window,'n_positive':n_positive,'vel_th_m_s':vel_th_m_s,
                   'mass_kg':mass_kg,'half_window_derivative':half_window_derivative}, f, ensure_ascii=False, indent=2)

    if csv_export:
        try:
            hoja3.to_csv(out_csv, index=False)
        except Exception:
            pass

    return out_xlsx, plot_path, out_json, (out_csv if csv_export else None)

# ---------------- batch driver ----------------

def process_batch(batch_dir: str, out_dir: str, **kwargs) -> List[Tuple[str, Tuple[str, str, str, Optional[str]]]]:
    os.makedirs(out_dir, exist_ok=True)
    entries = sorted([f for f in os.listdir(batch_dir) if f.lower().endswith('.xlsx')])
    results = []
    for name in entries:
        in_path = os.path.join(batch_dir, name)
        base = os.path.splitext(name)[0]
        out_prefix = os.path.join(out_dir, base + '_analysis_v3_3')
        out_xlsx = out_prefix + '.xlsx'
        # skip if already processed and newer than input
        if (not kwargs.get('overwrite', False)) and os.path.exists(out_xlsx):
            try:
                in_mtime  = os.path.getmtime(in_path)
                out_mtime = os.path.getmtime(out_xlsx)
                if out_mtime >= in_mtime:
                    print(f"SKIP (up-to-date): {name}")
                    continue
            except Exception:
                pass
        print(f"Processing: {name}")
        out = analyze_file(in_path, out_dir=out_dir, **kwargs)
        results.append((name, out))
    logger.info(f"Batch completo: procesados {len(results)} archivos")
    return results

# ---------------- main ----------------
if __name__ == '__main__':
    ap = argparse.ArgumentParser(description='STS v3.3: Acc Phases + Hoja3 ampliada + batch + CSV')
    ap.add_argument('--config', help='Archivo YAML de configuración')
    ap.add_argument('--show-config', action='store_true', help='Mostrar configuración y salir')
    ap.add_argument('--log-level', help='Nivel de logging (DEBUG/INFO/WARNING/ERROR/CRITICAL)')
    ap.add_argument('--log-dir', help='Directorio de logs (anula config)')
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument('--input', help='Ruta a un archivo .xlsx')
    g.add_argument('--batch', help='Carpeta con .xlsx a procesar')
    ap.add_argument('--out', default='data/output', help='Carpeta de salida (por defecto: data/output)')
    ap.add_argument('--sheet', default='Reducido')
    ap.add_argument('--time-col', default='time')
    ap.add_argument('--mass-kg', type=float, default=None)
    ap.add_argument('--window', type=int, default=30)
    ap.add_argument('--n-positive', type=int, default=30)
    ap.add_argument('--vel-th', type=float, default=0.1)
    ap.add_argument('--half-window-derivative', type=int, default=3)
    ap.add_argument('--csv', action='store_true', help='Exportar Hoja3 a CSV adicional')
    ap.add_argument('--overwrite', action='store_true', help='Reprocesar aunque ya exista el Excel de salida')
    args = ap.parse_args()

    # load configuration and logging
    config = load_config(args.config)
    if args.log_level:
        config.set('logging.level', args.log_level)
    log_dir = args.log_dir or config.get('log_directory')
    LoggerManager.configure(log_dir=log_dir)

    if args.show_config:
        print(json.dumps(config.to_dict(), indent=2))
        exit(0)

    # determine effective parameters (CLI overrides config)
    sheet = args.sheet or config.get('sheet_name', 'Reducido')
    timecol = args.time_col or config.get('time_col', 'time')
    mass = args.mass_kg if args.mass_kg is not None else config.get('mass_kg')
    window = args.window or config.get('window', 30)
    npos = args.n_positive or config.get('n_positive', 30)
    vth = args.vel_th or config.get('vel_th_m_s', 0.1)
    halfw = args.half_window_derivative or config.get('half_window_derivative', 3)

    if args.input:
        # single file
        out = analyze_file(
            input_path=args.input,
            sheet_name=sheet,
            time_col=timecol,
            mass_kg=mass,
            window=window,
            n_positive=npos,
            vel_th_m_s=vth,
            half_window_derivative=halfw,
            out_dir=args.out,
            csv_export=args.csv,
        )
        logger.info('LISTO (archivo único)')
        logger.info(f'Excel : {out[0]}')
        logger.info(f'Grafico: {out[1]}')
        logger.info(f'Params: {out[2]}')
        if args.csv:
            logger.info(f'CSV   : {out[3]}')
    else:
        # batch
        res = process_batch(
            batch_dir=args.batch,
            out_dir=args.out,
            sheet_name=sheet,
            time_col=timecol,
            mass_kg=mass,
            window=window,
            n_positive=npos,
            vel_th_m_s=vth,
            half_window_derivative=halfw,
            csv_export=args.csv,
            overwrite=args.overwrite,
        )
        logger.info('LISTO (batch)')
        logger.info(f'Procesados: {len(res)}')
