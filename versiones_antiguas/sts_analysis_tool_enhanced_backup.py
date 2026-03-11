#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Herramienta de análisis STS (Sit-to-Stand) - Versión Enhanced.

Features:
- Detección exacta de fases estilo Excel
- Soporte para Hip Z (desplazamiento y velocidad)
- Soporte para EMG (electromiografía)
- Soporte para CoP SD (estabilidad Centro de Presión)
- "Acc Phases" (marcadores acumulativos de fase)
- Metadatos Noraxon (código, altura, peso, etc.)
- Hoja3 ampliada (Kinematic & Forces like)

Requiere: numpy, pandas, matplotlib, openpyxl
"""

import argparse
import os
from typing import Optional, Tuple, List

import numpy as np
import pandas as pd

from src.utils import centered_slope, cumulative_trapezoid, detect_column
from src.detection import (
    compute_vector_displacements,
    compute_windows,
    detect_conc_starts,
    detect_peaks,
    detect_ecc_ends,
    pair_repetitions,
    compute_phase_events,
    compute_acc_phases,
)
from src.metrics import RepResult, compute_metrics, mark_ok_repetitions
from src.plotting import generate_plots
from src.export import create_sheet1_variables, export_to_excel, export_to_json, export_advanced_sheet3
from src.advanced_metrics import PhaseComputer, detect_emg_columns, detect_cop_columns
from src.metadata import read_metadata


def run_tool_enhanced(
    file_path: str,
    sheet_name: str = 'Reducido',
    time_col: str = 'time',
    disp_col: Optional[str] = None,
    mass_kg: Optional[float] = None,
    window: int = 30,
    n_positive: int = 30,
    vel_th_m_s: float = 0.1,
    ok_th: float = 0.85,
    half_window_derivative: int = 3,
    out_prefix: Optional[str] = None,
    out_dir: Optional[str] = None,
    csv_export: bool = False,
    make_plot: bool = True,
    per_rep_plots: bool = True,
    close_last_seated_at_end: bool = True,
) -> Tuple[str, Optional[str], str, Optional[str]]:
    """
    Ejecuta análisis STS completo con funcionalidades avanzadas.
    
    Args:
        file_path: Ruta del archivo Excel.
        sheet_name: Hoja a analizar.
        time_col: Columna de tiempo.
        disp_col: Columna de desplazamiento (autodetecta si es None).
        mass_kg: Masa corporal para calcular potencia.
        window: Tamaño de ventana de detección.
        n_positive: N de celdas positivas/negativas para eventos.
        vel_th_m_s: Umbral de velocidad.
        ok_th: Umbral relativo de amplitud.
        half_window_derivative: Semiventana para derivadas.
        out_prefix: Prefijo de salida (override out_dir si se especifica).
        out_dir: Directorio de salida (auto-genera nombre si se usa con batch).
        csv_export: Exportar Hoja3 a CSV.
        make_plot: Generar gráfico general.
        per_rep_plots: Generar gráficos por repetición.
        close_last_seated_at_end: Cerrar fase "Sentado" al final.
        
    Returns:
        (excel_path, plot_path, json_path, csv_path or None)
    """
    # --------
    # 1. LEER DATOS
    # --------
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    if time_col not in df.columns:
        raise ValueError(f'No se encontró la columna de tiempo "{time_col}".')
    
    t = df[time_col].astype(float).to_numpy()
    if len(t) < 3:
        raise ValueError('No hay suficientes muestras para analizar.')
    
    dt = float(pd.Series(np.diff(t)).median())
    n = len(t)
    
    # Detectar columna de desplazamiento BCM Z
    if disp_col is None:
        disp_col = detect_column(
            df,
            candidates=['Disp_BCM_Z', 'BCM_Z', 'Body center of mass-z (mm)'],
            contains_all=['body center of mass', 'z', '(mm)']
        )
        if disp_col is None:
            for c in df.columns:
                cl = str(c).lower()
                if ('body center of mass' in cl) and ('z' in cl) and ('mm' in cl):
                    disp_col = c
                    break
    
    if disp_col is None:
        raise ValueError('No se encontró la columna de desplazamiento BCM en eje Z.')
    
    z_mm = df[disp_col].astype(float).to_numpy()
    z_m = z_mm / 1000.0
    
    # --------
    # 2. CALCULAR VARIABLES DERIVADAS
    # --------
    vel_m_s = centered_slope(z_m, dt, half_window_derivative)
    acc_m_s2 = centered_slope(vel_m_s, dt, half_window_derivative)
    
    force_est_N = np.full_like(vel_m_s, np.nan)
    if mass_kg is not None:
        force_est_N[:] = mass_kg * 9.80665
        power_W = force_est_N * vel_m_s
    else:
        power_W = np.full_like(vel_m_s, np.nan)
    
    work_cum_J = cumulative_trapezoid(power_W, dt)
    
    # --------
    # 3. DETECCIÓN DE EVENTOS
    # --------
    vector_disp = compute_vector_displacements(z_mm)
    future_sum, previous_sum = compute_windows(vector_disp, window)
    
    vel_conc_flag = np.where(vel_m_s > vel_th_m_s, 1, np.nan)
    vel_ecc_flag = np.where(vel_m_s < -vel_th_m_s, 1, np.nan)
    
    conc_start = detect_conc_starts(future_sum, vel_conc_flag, n_positive)
    conc_exc = detect_peaks(z_mm, window)
    ecc_end = detect_ecc_ends(previous_sum, vel_ecc_flag, n_positive)
    
    conc_event, conc_graph, ecc_event, ecc_graph, any_phase_event = compute_phase_events(
        conc_start, conc_exc, ecc_end
    )
    
    # --------
    # 4. EMPAREJAR REPETICIONES
    # --------
    idx_starts = list(np.where(conc_start == 10000)[0])
    idx_peaks = list(np.where(conc_exc == 10000)[0])
    idx_ends = list(np.where(ecc_end == 10000)[0])
    
    reps_data, phase_id, phase_label = pair_repetitions(idx_starts, idx_peaks, idx_ends, n)
    
    # --------
    # 5. CALCULAR ACC_PHASES Y MARKER_TO_TIME
    # --------
    acc_phases, marker_to_time, rep_events = compute_acc_phases(conc_start, conc_exc, ecc_end, t)
    
    # --------
    # 6. CALCULAR MÉTRICAS POR REPETICIÓN
    # --------
    reps: List[RepResult] = []
    for rep_num, (s, p, e) in enumerate(reps_data, start=1):
        next_s = idx_starts[rep_num] if rep_num < len(idx_starts) else n - 1
        rep = compute_metrics(
            rep_num, s, p, e, next_s, n,
            z_mm, vel_m_s, acc_m_s2, power_W,
            t, dt
        )
        reps.append(rep)
    
    mark_ok_repetitions(reps, ok_th)
    reps_df = pd.DataFrame([r.__dict__ for r in reps])
    
    # --------
    # 7. CREAR DATAFRAME HOJA1
    # --------
    rep_id = np.full(n, np.nan)
    for rr in reps:
        s, p, e = rr.idx_conc_start, rr.idx_peak, rr.idx_ecc_end
        if s is not None and p is not None:
            rep_id[s:p + 1] = rr.rep
        if p is not None and e is not None:
            rep_id[p:e + 1] = rr.rep
        if e is not None:
            next_starts = [x for x in idx_starts if x > e]
            ns = next_starts[0] if next_starts else n
            rep_id[e + 1:ns] = rr.rep
    
    sheet1 = create_sheet1_variables(
        df, z_mm, vel_m_s, acc_m_s2, force_est_N, power_W, work_cum_J,
        vector_disp, future_sum, previous_sum,
        vel_conc_flag, vel_ecc_flag,
        conc_start, conc_exc, ecc_end,
        conc_event, conc_graph, ecc_event, ecc_graph, any_phase_event,
        phase_id, phase_label, rep_id
    )
    
    # Agregar Acc Phases a Hoja1
    sheet1['Acc Phases'] = acc_phases
    
    # --------
    # 8. CREAR HOJA3 AMPLIADA (KINEMATIC & FORCES)
    # --------
    metadata = read_metadata(file_path, mass_kg)
    pc = PhaseComputer(df, time_col)
    
    emg_cols = detect_emg_columns(df)
    cop_cols = detect_cop_columns(df)
    
    # Detectar Hip Z
    disp_hipz = next((c for c in df.columns if 'hip rt-z' in str(c).lower() 
                     or 'hip' in str(c).lower() and 'z' in str(c).lower()), None)
    
    hoja3_rows = []
    for rep_event in rep_events:
        r = rep_event.rep
        
        for phase_name in ['Concéntrica', 'Excéntrica', 'Sentado']:
            # Calcular tiempos de fase
            m1 = (r - 1) * 3 + 1
            m2 = m1 + 1
            m3 = m1 + 2
            
            if phase_name == 'Concéntrica':
                sm, em = m1, m2
            elif phase_name == 'Excéntrica':
                sm, em = m2, m3
            else:  # Sentado
                sm, em = m3, m3 + 1
            
            if sm not in marker_to_time:
                continue
            
            t0 = marker_to_time[sm]
            t1 = marker_to_time[em] if em in marker_to_time else (float(t[-1]) if close_last_seated_at_end else t0)
            
            row = {
                'Código': metadata['Codigo'],
                'Altura': metadata['Altura'],
                'Peso_kg': metadata['Peso_kg'],
                'Altura de la silla': metadata['Altura_silla'],
                'Test': metadata['Test'],
                'Fecha Test': metadata['Fecha_Test'],
                'Repeticion': r,
                'Fase': phase_name,
                'Inicio_Tiempo': t0,
                'Final_Tiempo': t1,
                'Tiempo fase': max(0.0, t1 - t0),
            }
            
            # BCM Z (siempre)
            mx, mn, rg = pc.range_mm_to_m(disp_col, t0, t1)
            row['BCM_Z_Max (m)'] = mx
            row['BCM_Z_Min (m)'] = mn
            row['BCM_Z_Range (m)'] = rg
            
            mV, xV, minV = pc.stats(None, vel_m_s, t0, t1)
            row['BCM_Z_Vel_Mean (m/s)'] = mV
            row['BCM_Z_Vel_Max (m/s)'] = xV
            row['BCM_Z_Vel_Min (m/s)'] = minV
            
            mP, xP, work = pc.power_work(None, None, metadata['Peso_kg'], t0, t1)
            row['BCM_Z_Power_Mean (W)'] = mP
            row['BCM_Z_Power_Max (W)'] = xP
            row['BCM_Z_Work (J)'] = work
            
            # Hip Z
            if disp_hipz is not None:
                hmx, hmn, hrg = pc.range_mm_to_m(disp_hipz, t0, t1)
                row['HIP_Z_Max (m)'] = hmx
                row['HIP_Z_Min (m)'] = hmn
                row['HIP_Z_Range (m)'] = hrg
                
                # Derivar velocidad de Hip
                hip_mm = pd.to_numeric(df[disp_hipz], errors='coerce').to_numpy(dtype=float)
                hip_m = hip_mm / 1000.0
                hip_vel = centered_slope(hip_m, dt, half_window_derivative)
                hVm, hVx, hVn = pc.stats(None, hip_vel, t0, t1)
                row['HIP_Z_Vel_Mean (m/s)'] = hVm
                row['HIP_Z_Vel_Max (m/s)'] = hVx
                row['HIP_Z_Vel_Min (m/s)'] = hVn
            
            # EMG medios
            for emg in emg_cols:
                row[f'EMG_Mean_{emg}'] = pc.mean(emg, t0, t1)
            
            # CoP SD
            for key, col in cop_cols.items():
                if col is not None:
                    if 'SD' in col:
                        row[f'CoP_{key}'] = pc.mean(col, t0, t1)
                    else:
                        row[f'CoP_{key}_SD'] = pc.stdev(col, t0, t1)
            
            hoja3_rows.append(row)
    
    hoja3_df = pd.DataFrame(hoja3_rows)
    
    # --------
    # 9. GENERAR GRÁFICOS Y CONFIGURAR RUTAS DE SALIDA
    # --------
    # Generar out_prefix a partir de out_dir si es necesario
    if not out_prefix:
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            out_prefix = os.path.join(out_dir, base_name + '_analysis_enhanced')
        else:
            out_prefix = os.path.splitext(file_path)[0] + '_analysis_enhanced'
    
    plot_path, rep_dir = generate_plots(
        t, z_mm, vel_m_s,
        idx_starts, idx_peaks, idx_ends,
        reps, out_prefix,
        make_plot, per_rep_plots
    )
    
    # --------
    # 10. EXPORTAR DATOS
    # --------
    params = {
        'input_file': file_path,
        'sheet_name': sheet_name,
        'time_col': time_col,
        'disp_col': disp_col,
        'dt_s': dt,
        'window': window,
        'n_positive': n_positive,
        'vel_th_m_s': vel_th_m_s,
        'ok_th': ok_th,
        'mass_kg': mass_kg,
        'half_window_derivative': half_window_derivative,
        'n_conc_start': len(idx_starts),
        'n_peaks': len(idx_peaks),
        'n_ecc_end': len(idx_ends),
        'n_reps': len(reps),
        'features': ['BCM_Z', 'Hip_Z', 'EMG', 'CoP_SD', 'Acc_Phases'] if any([disp_hipz, emg_cols, cop_cols]) else ['BCM_Z', 'Acc_Phases'],
    }
    
    # Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    
    out_excel = out_prefix + '.xlsx'
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Hoja1_Variables'
    
    for j, col in enumerate(sheet1.columns, start=1):
        cell = ws1.cell(row=1, column=j, value=str(col))
        cell.font = Font(bold=True)
    for i, (_, row) in enumerate(sheet1.iterrows(), start=2):
        for j, val in enumerate(row, start=1):
            if isinstance(val, (np.floating, float)) and (np.isnan(val) or np.isinf(val)):
                val = None
            ws1.cell(row=i, column=j, value=val)
    
    for j, col in enumerate(sheet1.columns, start=1):
        ws1.column_dimensions[get_column_letter(j)].width = min(max(len(str(col)) + 2, 12), 28)
    
    # Hoja2
    ws2 = wb.create_sheet('Hoja2_Parametros_Grafica')
    param_list = [
        ('Archivo entrada', params.get('input_file')),
        ('Hoja analizada', params.get('sheet_name')),
        ('Columna tiempo', params.get('time_col')),
        ('Columna desplazamiento BCM Z', params.get('disp_col')),
        ('dt (s)', params.get('dt_s')),
        ('Ventana vectores', params.get('window')),
        ('N celdas positivas/negativas', params.get('n_positive')),
        ('Umbral velocidad (m/s)', params.get('vel_th_m_s')),
        ('Umbral amplitud OK', params.get('ok_th')),
        ('Masa (kg)', params.get('mass_kg')),
        ('Derivada centrada (half-window)', params.get('half_window_derivative')),
        ('N starts detectados', params.get('n_conc_start')),
        ('N picos detectados', params.get('n_peaks')),
        ('N ecc_end detectados', params.get('n_ecc_end')),
        ('N repeticiones emparejadas', params.get('n_reps')),
        ('Features avanzadas', ', '.join(params.get('features', []))),
    ]
    
    for r, (k, v) in enumerate(param_list, start=1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    
    if plot_path and os.path.exists(plot_path):
        img = XLImage(plot_path)
        img.width = 1100
        img.height = 550
        ws2.add_image(img, 'D2')
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 18
    
    # Hoja3 ampliada
    export_advanced_sheet3(wb, hoja3_df)
    
    wb.save(out_excel)
    
    # JSON
    out_json = out_prefix + '_params.json'
    export_to_json(params, out_json)
    
    # CSV (Hoja3 si se solicita)
    out_csv = None
    if csv_export:
        try:
            out_csv = out_prefix + '_hoja3.csv'
            hoja3_df.to_csv(out_csv, index=False)
        except Exception as e:
            print(f"Warning: No se pudo exportar CSV: {e}")
    
    return out_excel, plot_path, out_json, out_csv


def analyze_file(
    input_path: str,
    out_dir: str = 'data/output',
    **kwargs
) -> Tuple[str, Optional[str], str, Optional[str]]:
    """
    Wrapper para procesar un archivo con configuración de directorio de salida.
    
    Args:
        input_path: Ruta del archivo Excel a analizar
        out_dir: Directorio de salida
        **kwargs: Argumentos adicionales para run_tool_enhanced
        
    Returns:
        (excel_path, plot_path, json_path, csv_path)
    """
    os.makedirs(out_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    out_prefix = os.path.join(out_dir, base_name + '_analysis_enhanced')
    
    return run_tool_enhanced(
        file_path=input_path,
        out_prefix=out_prefix,
        **kwargs
    )


def process_batch(
    batch_dir: str,
    out_dir: str = 'data/output',
    **kwargs
) -> List[Tuple[str, Tuple]]:
    """
    Procesa batch de archivos Excel en una carpeta.
    
    Args:
        batch_dir: Carpeta con archivos .xlsx
        out_dir: Directorio de salida
        **kwargs: Argumentos adicionales para run_tool_enhanced
        
    Returns:
        Lista de (nombre_archivo, (excel, plot, json, csv))
    """
    os.makedirs(out_dir, exist_ok=True)
    entries = sorted([f for f in os.listdir(batch_dir) if f.lower().endswith('.xlsx')])
    results = []
    
    for name in entries:
        in_path = os.path.join(batch_dir, name)
        base = os.path.splitext(name)[0]
        out_prefix = os.path.join(out_dir, base + '_analysis_enhanced')
        out_xlsx = out_prefix + '.xlsx'
        
        # Skip if already processed and newer than input (unless overwrite=True)
        if not kwargs.get('overwrite', False) and os.path.exists(out_xlsx):
            try:
                in_mtime = os.path.getmtime(in_path)
                out_mtime = os.path.getmtime(out_xlsx)
                if out_mtime >= in_mtime:
                    print(f"SKIP (up-to-date): {name}")
                    continue
            except Exception:
                pass
        
        print(f"Processing: {name}")
        try:
            out = run_tool_enhanced(
                file_path=in_path,
                out_prefix=out_prefix,
                **kwargs
            )
            results.append((name, out))
        except Exception as e:
            print(f"  ERROR: {e}")
            results.append((name, (None, None, None, None)))
    
    return results


def main():
    """Interfaz principal con soporte para modo individual y batch."""
    parser = argparse.ArgumentParser(
        description='Análisis STS Enhanced v2.0: Hip Z + EMG + CoP_SD + Batch'
    )
    
    # Grupo mutuamente exclusivo: --input vs --batch
    g = parser.add_mutually_exclusive_group(required=True)
    g.add_argument('--input', help='Ruta a un archivo .xlsx individual')
    g.add_argument('--batch', help='Carpeta con .xlsx a procesar en lote')
    
    # Argumentos comunes
    parser.add_argument('--out', default='data/output', help='Carpeta de salida (default: data/output)')
    parser.add_argument('--sheet', default='Reducido', help='Hoja a procesar')
    parser.add_argument('--time-col', default='time', help='Columna de tiempo')
    parser.add_argument('--disp-col', default=None, help='Columna BCM Z (auto-detect si omite)')
    parser.add_argument('--mass-kg', type=float, default=None, help='Masa corporal (kg)')
    parser.add_argument('--window', type=int, default=30, help='Ventana de detección')
    parser.add_argument('--n-positive', type=int, default=30, help='N celdas positivas/negativas')
    parser.add_argument('--vel-th', type=float, default=0.1, help='Umbral velocidad (m/s)')
    parser.add_argument('--ok-th', type=float, default=0.85, help='Umbral amplitud OK')
    parser.add_argument('--half-window-derivative', type=int, default=3, help='Semiventana derivadas')
    parser.add_argument('--csv', action='store_true', help='Exportar Hoja3 a CSV adicional')
    parser.add_argument('--no-plot', action='store_true', help='No generar gráfico general')
    parser.add_argument('--no-per-rep', action='store_true', help='No generar gráficos por repetición')
    parser.add_argument('--overwrite', action='store_true', help='Reprocesar aunque exista output (batch)')
    
    args = parser.parse_args()
    
    # Argumentos comunes para run_tool_enhanced
    common_kwargs = {
        'sheet_name': args.sheet,
        'time_col': args.time_col,
        'disp_col': args.disp_col,
        'mass_kg': args.mass_kg,
        'window': args.window,
        'n_positive': args.n_positive,
        'vel_th_m_s': args.vel_th,
        'ok_th': args.ok_th,
        'half_window_derivative': args.half_window_derivative,
        'csv_export': args.csv,
        'make_plot': (not args.no_plot),
        'per_rep_plots': (not args.no_per_rep),
    }
    
    if args.input:
        # Modo individual
        print(f"\n Processing individual file: {args.input}")
        try:
            excel, plot, json, csv = run_tool_enhanced(
                file_path=args.input,
                out_dir=args.out,
                **common_kwargs
            )
            print('\n√ ANALYSIS COMPLETE')
            print(f'  Excel:  {excel}')
            print(f'  Plot:   {plot}')
            print(f'  Params: {json}')
            if csv:
                print(f'  CSV:    {csv}')
        except Exception as e:
            print(f'\n ERROR: {e}')
            
    else:
        # Modo batch
        print(f"\n Batch processing: {args.batch}")
        print(f"   Output dir: {args.out}")
        results = process_batch(
            batch_dir=args.batch,
            out_dir=args.out,
            overwrite=args.overwrite,
            **common_kwargs
        )
        
        print(f'\n√ BATCH COMPLETE')
        print(f'  Total processed: {len(results)}')
        success = sum(1 for _, (e, _, _, _) in results if e is not None)
        print(f'  Success: {success}')
        print(f'  Failed: {len(results) - success}')



        description='Análisis STS Enhanced: detección exacta + Hip Z + EMG + CoP SD'
    )
    parser.add_argument('--input', required=True, help='Ruta del archivo .xlsx de entrada')
    parser.add_argument('--sheet', default='Reducido', help='Hoja a procesar')
    parser.add_argument('--time-col', default='time', help='Nombre de la columna de tiempo')
    parser.add_argument('--disp-col', default=None, help='Nombre exacto de la columna de desplazamiento BCM Z')
    parser.add_argument('--mass-kg', type=float, default=None, help='Masa corporal (kg)')
    parser.add_argument('--window', type=int, default=30, help='Ventana de detección')
    parser.add_argument('--n-positive', type=int, default=30, help='N celdas positivas/negativas')
    parser.add_argument('--vel-th', type=float, default=0.1, help='Umbral de velocidad (m/s)')
    parser.add_argument('--ok-th', type=float, default=0.85, help='Umbral relativo de amplitud')
    parser.add_argument('--half-window-derivative', type=int, default=3, help='Semiventana para derivadas')
    parser.add_argument('--no-plot', action='store_true', help='No generar gráfico general')
    parser.add_argument('--no-per-rep', action='store_true', help='No generar gráficos por repetición')
    parser.add_argument('--out-prefix', default=None, help='Prefijo de salida')
    args = parser.parse_args()
    
    excel, plot, params = run_tool_enhanced(
        file_path=args.input,
        sheet_name=args.sheet,
        time_col=args.time_col,
        disp_col=args.disp_col,
        mass_kg=args.mass_kg,
        window=args.window,
        n_positive=args.n_positive,
        vel_th_m_s=args.vel_th,
        ok_th=args.ok_th,
        half_window_derivative=args.half_window_derivative,
        out_prefix=args.out_prefix,
        make_plot=(not args.no_plot),
        per_rep_plots=(not args.no_per_rep),
    )
    
    print('\n✓ ANALYSIS COMPLETE')
    print(f'Excel: {excel}')
    print(f'Plot: {plot}')
    print(f'Params: {params}')


if __name__ == '__main__':
    main()
