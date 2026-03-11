"""
Exportación a Excel y JSON.
"""

import os
import json
from typing import List, Optional
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from src.metrics import RepResult


def create_sheet1_variables(
    df_original: pd.DataFrame,
    z_mm: np.ndarray,
    vel_m_s: np.ndarray,
    acc_m_s2: np.ndarray,
    force_est_N: np.ndarray,
    power_W: np.ndarray,
    work_cum_J: np.ndarray,
    vector_disp: np.ndarray,
    future_sum: np.ndarray,
    previous_sum: np.ndarray,
    vel_conc_flag: np.ndarray,
    vel_ecc_flag: np.ndarray,
    conc_start: np.ndarray,
    conc_exc: np.ndarray,
    ecc_end: np.ndarray,
    conc_event: np.ndarray,
    conc_graph: np.ndarray,
    ecc_event: np.ndarray,
    ecc_graph: np.ndarray,
    any_phase_event: np.ndarray,
    phase_id: np.ndarray,
    phase_label: np.ndarray,
    rep_id: np.ndarray,
) -> pd.DataFrame:
    """
    Crea DataFrame para Hoja1 (todas las variables).
    
    Returns:
        DataFrame con todas las columnas.
    """
    sheet1 = df_original.copy()
    
    sheet1['Disp_BCM_Z_mm'] = z_mm
    sheet1['Vel_BCM_Z_m_s'] = vel_m_s
    sheet1['Acc_BCM_Z_m_s2'] = acc_m_s2
    sheet1['Force_Estimate_BCM_Z_N'] = force_est_N
    sheet1['Power_Estimated_BCM_Z_W'] = power_W
    sheet1['Work_Cumulative_BCM_Z_J'] = work_cum_J
    
    sheet1['Cod Vector Disp'] = vector_disp
    sheet1['Subs Cells_Displ'] = future_sum
    sheet1['Previous Cells_Disp'] = previous_sum
    sheet1['Vel Conc>Vel Umbral'] = vel_conc_flag
    sheet1['Vel Ecc < Vel Umbral'] = vel_ecc_flag
    sheet1['Conc Start'] = np.where(conc_start == 10000, 10000, 0)
    sheet1['Conc-Exc'] = np.where(conc_exc == 10000, 10000, 0)
    sheet1['Ecc End'] = np.where(ecc_end == 10000, 10000, 0)
    sheet1['Conc Phases'] = conc_event
    sheet1['Conc Graph'] = conc_graph
    sheet1['Ecc Phases'] = ecc_event
    sheet1['Ecc Graph'] = ecc_graph
    sheet1['Phases'] = any_phase_event
    sheet1['Phase_ID'] = phase_id
    sheet1['Phase_Label'] = phase_label
    sheet1['Rep_ID'] = rep_id
    
    return sheet1


def export_to_excel(
    sheet1: pd.DataFrame,
    reps_df: pd.DataFrame,
    params: dict,
    plot_path: Optional[str],
    out_excel: str
) -> None:
    """
    Exporta datos a archivo Excel con 3 hojas.
    
    Args:
        sheet1: DataFrame de variables.
        reps_df: DataFrame de resultados de repeticiones.
        params: Diccionario de parámetros.
        plot_path: Ruta del gráfico (opcional).
        out_excel: Ruta de salida del Excel.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Hoja1_Variables'
    
    # Hoja 1: Variables
    for j, col in enumerate(sheet1.columns, start=1):
        cell = ws1.cell(row=1, column=j, value=str(col))
        cell.font = Font(bold=True)
    for i, (_, row) in enumerate(sheet1.iterrows(), start=2):
        for j, val in enumerate(row, start=1):
            if isinstance(val, (np.floating, float)) and (np.isnan(val) or np.isinf(val)):
                val = None
            ws1.cell(row=i, column=j, value=val)
    
    for j, col in enumerate(sheet1.columns, start=1):
        ws1.column_dimensions[get_column_letter(j)].width = min(max(len(str(col)) + 2, 12), 28)
    
    # Hoja 2: Parámetros + Gráfica
    ws2 = wb.create_sheet('Hoja2_Parametros_Grafica')
    param_list = [
        ('Archivo entrada', params.get('input_file')),
        ('Hoja analizada', params.get('sheet_name')),
        ('Columna tiempo', params.get('time_col')),
        ('Columna desplazamiento BCM Z', params.get('disp_col')),
        ('dt (s)', params.get('dt_s')),
        ('Ventana vectores', params.get('window')),
        ('N celdas positivas/negativas', params.get('n_positive')),
        ('Umbral velocidad (m/s)', params.get('vel_th_m_s')),
        ('Umbral amplitud OK', params.get('ok_th')),
        ('Masa (kg)', params.get('mass_kg')),
        ('Derivada centrada (half-window)', params.get('half_window_derivative')),
        ('N starts detectados', params.get('n_conc_start')),
        ('N picos detectados', params.get('n_peaks')),
        ('N ecc_end detectados', params.get('n_ecc_end')),
        ('N repeticiones emparejadas', params.get('n_reps')),
    ]
    
    for r, (k, v) in enumerate(param_list, start=1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    
    # Resumen global
    start_row_summary = len(param_list) + 3
    ws2.cell(row=start_row_summary, column=1, value='Resumen global').font = Font(bold=True)
    
    if len(reps_df):
        summary_items = [
            ('Amplitud máxima subida (mm)', float(np.nanmax(reps_df['amp_up_mm'])) if reps_df['amp_up_mm'].notna().any() else None),
            ('Amplitud media subida (mm)', float(np.nanmean(reps_df['amp_up_mm'])) if reps_df['amp_up_mm'].notna().any() else None),
            ('Duración media levantarse (s)', float(np.nanmean(reps_df['dur_up_s'])) if reps_df['dur_up_s'].notna().any() else None),
            ('Duración media sentarse (s)', float(np.nanmean(reps_df['dur_down_s'])) if reps_df['dur_down_s'].notna().any() else None),
            ('Duración media sentado (s)', float(np.nanmean(reps_df['dur_seated_after_s'])) if reps_df['dur_seated_after_s'].notna().any() else None),
            ('N OK amplitud', int(np.nansum(reps_df['ok_up'])) if reps_df['ok_up'].notna().any() else None),
        ]
        for k, (kk, vv) in enumerate(summary_items, start=start_row_summary + 1):
            ws2.cell(row=k, column=1, value=kk).font = Font(bold=False)
            ws2.cell(row=k, column=2, value=vv)
    
    # Insertar gráfico
    if plot_path and os.path.exists(plot_path):
        img = XLImage(plot_path)
        img.width = 1100
        img.height = 550
        ws2.add_image(img, 'D2')
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 18
    
    # Hoja 3: Resultados por repetición
    ws3 = wb.create_sheet('Hoja3_Resultados_Repeticion')
    if not reps_df.empty:
        for j, col in enumerate(reps_df.columns, start=1):
            ws3.cell(row=1, column=j, value=col).font = Font(bold=True)
        for i, (_, row) in enumerate(reps_df.iterrows(), start=2):
            for j, val in enumerate(row, start=1):
                if isinstance(val, (np.floating, float)) and (np.isnan(val) or np.isinf(val)):
                    val = None
                ws3.cell(row=i, column=j, value=val)
        for j, col in enumerate(reps_df.columns, start=1):
            ws3.column_dimensions[get_column_letter(j)].width = min(max(len(str(col)) + 2, 14), 24)
    else:
        ws3['A1'] = 'No se detectaron repeticiones.'
    
    wb.save(out_excel)


def export_to_json(params: dict, out_json: str) -> None:
    """
    Exporta parámetros a JSON.
    
    Args:
        params: Diccionario de parámetros.
        out_json: Ruta de salida.
    """
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump(params, f, ensure_ascii=False, indent=2)
