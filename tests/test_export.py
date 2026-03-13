"""
Tests for src/export.py module.

Tests cover:
- Excel export functionality (create_sheet1_variables, export_to_excel)
- JSON export functionality (export_to_json)
- Advanced sheet export (export_advanced_sheet3)
- Data formatting and validation
- Error handling for missing files/data
"""

import pytest
import pandas as pd
import numpy as np
import json
import os
import tempfile
from pathlib import Path
from openpyxl import Workbook, load_workbook
from src.export import (
    create_sheet1_variables,
    export_to_excel,
    export_to_json,
    export_advanced_sheet3,
)
from src.metrics import RepResult


class TestCreateSheet1Variables:
    """Tests for create_sheet1_variables function."""

    def create_test_data(self):
        """Create test DataFrame and arrays."""
        df = pd.DataFrame({
            'Tiempo': np.arange(0, 1.0, 0.01),
            'BCM Z': 600 + np.random.randn(100),
            'Velocidad': np.random.randn(100) * 0.1,
            'Aceleración': np.random.randn(100) * 0.1,
        })

        # Create dummy arrays for all required parameters
        n_samples = len(df)
        z_mm = df['BCM Z'].values
        vel_m_s = df['Velocidad'].values
        acc_m_s2 = df['Aceleración'].values
        force_est_N = np.random.randn(n_samples) * 100
        power_W = np.random.randn(n_samples) * 50
        work_cum_J = np.cumsum(np.abs(power_W)) * 0.01
        vector_disp = np.random.randn(n_samples)
        future_sum = np.random.randn(n_samples)
        previous_sum = np.random.randn(n_samples)
        vel_conc_flag = np.random.randint(0, 2, n_samples)
        vel_ecc_flag = np.random.randint(0, 2, n_samples)
        conc_start = np.random.randint(0, 2, n_samples)
        conc_exc = np.random.randint(0, 2, n_samples)
        ecc_end = np.random.randint(0, 2, n_samples)
        conc_event = np.random.randint(0, 2, n_samples)
        conc_graph = np.random.randint(0, 2, n_samples)
        ecc_event = np.random.randint(0, 2, n_samples)
        ecc_graph = np.random.randint(0, 2, n_samples)
        any_phase_event = np.random.randint(0, 2, n_samples)
        phase_id = np.random.randint(0, 4, n_samples)
        phase_label = np.array(['Seated'] * 25 + ['Conc'] * 25 + ['Ecc'] * 25 + ['Seated'] * 25)
        rep_id = np.repeat(range(1, 5), 25)

        return {
            'df_original': df,
            'z_mm': z_mm,
            'vel_m_s': vel_m_s,
            'acc_m_s2': acc_m_s2,
            'force_est_N': force_est_N,
            'power_W': power_W,
            'work_cum_J': work_cum_J,
            'vector_disp': vector_disp,
            'future_sum': future_sum,
            'previous_sum': previous_sum,
            'vel_conc_flag': vel_conc_flag,
            'vel_ecc_flag': vel_ecc_flag,
            'conc_start': conc_start,
            'conc_exc': conc_exc,
            'ecc_end': ecc_end,
            'conc_event': conc_event,
            'conc_graph': conc_graph,
            'ecc_event': ecc_event,
            'ecc_graph': ecc_graph,
            'any_phase_event': any_phase_event,
            'phase_id': phase_id,
            'phase_label': phase_label,
            'rep_id': rep_id,
        }

    def test_creates_sheet1_dataframe(self):
        """Should create DataFrame with all required columns."""
        data = self.create_test_data()

        result = create_sheet1_variables(**data)

        assert isinstance(result, pd.DataFrame)
        assert len(result) == len(data['df_original'])

        # Check that original columns are preserved
        for col in data['df_original'].columns:
            assert col in result.columns

        # Check that new columns are added
        expected_new_cols = [
            'Disp_BCM_Z_mm', 'Vel_BCM_Z_m_s', 'Acc_BCM_Z_m_s2', 'Force_Estimate_BCM_Z_N', 'Power_Estimated_BCM_Z_W', 'Work_Cumulative_BCM_Z_J',
            'Cod Vector Disp', 'Subs Cells_Displ', 'Previous Cells_Disp', 'Vel Conc>Vel Umbral',
            'Vel Ecc < Vel Umbral', 'Conc Start', 'Conc-Exc', 'Ecc End', 'Conc Phases',
            'Conc Graph', 'Ecc Phases', 'Ecc Graph', 'Phases', 'Phase_ID', 'Phase_Label', 'Rep_ID'
        ]

        for col in expected_new_cols:
            assert col in result.columns

    def test_preserves_original_data(self):
        """Should preserve original DataFrame data."""
        data = self.create_test_data()
        original_values = data['df_original'].copy()

        result = create_sheet1_variables(**data)

        # Check that original columns have same values
        for col in original_values.columns:
            pd.testing.assert_series_equal(
                result[col], original_values[col], check_names=False
            )

    def test_handles_nan_values(self):
        """Should handle NaN values appropriately."""
        data = self.create_test_data()

        # Add some NaN values
        data['power_W'][10:20] = np.nan
        data['work_cum_J'][10:20] = np.nan

        result = create_sheet1_variables(**data)

        # Should not crash and should preserve NaN
        assert result['Power_Estimated_BCM_Z_W'].isna().sum() > 0
        assert result['Work_Cumulative_BCM_Z_J'].isna().sum() > 0


class TestExportToExcel:
    """Tests for export_to_excel function."""

    def create_test_data(self):
        """Create test data for Excel export."""
        df = pd.DataFrame({
            'Tiempo': np.arange(0, 1.0, 0.01),
            'BCM Z': 600 + np.random.randn(100),
            'Velocidad': np.random.randn(100) * 0.1,
        })

        # Create mock RepResult objects
        reps = [
            RepResult(
                rep=1,
                idx_conc_start=100,
                idx_peak=200,
                idx_ecc_end=300,
                t_conc_start_s=1.0,
                t_peak_s=2.0,
                t_ecc_end_s=3.0,
                amp_up_mm=50.0,
                amp_down_mm=45.0,
                dur_up_s=1.5,
                dur_down_s=1.2,
                dur_seated_after_s=0.5,
                vmax_up_m_s=0.8,
                vmin_down_m_s=-0.7,
                t_to_vmax_up_s=0.8,
                t_to_vmin_down_s=0.6,
                t_pos_acc_up_s=0.4,
                t_neg_acc_down_s=0.3,
                pmax_up_W=150.0,
                pmin_down_W=-120.0,
                work_up_J=25.0,
                work_down_J=20.0,
                ok_up=1
            ),
            RepResult(
                rep=2,
                idx_conc_start=400,
                idx_peak=500,
                idx_ecc_end=600,
                t_conc_start_s=4.0,
                t_peak_s=5.0,
                t_ecc_end_s=6.0,
                amp_up_mm=48.0,
                amp_down_mm=47.0,
                dur_up_s=1.4,
                dur_down_s=1.3,
                dur_seated_after_s=0.6,
                vmax_up_m_s=0.75,
                vmin_down_m_s=-0.65,
                t_to_vmax_up_s=0.7,
                t_to_vmin_down_s=0.5,
                t_pos_acc_up_s=0.35,
                t_neg_acc_down_s=0.25,
                pmax_up_W=140.0,
                pmin_down_W=-110.0,
                work_up_J=22.0,
                work_down_J=18.0,
                ok_up=1
            )
        ]

        return df, reps

    def test_exports_excel_file(self):
        """Should create Excel file with correct sheets."""
        df, reps = self.create_test_data()

        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = os.path.join(temp_dir, 'test.xlsx')
            plot_path = None  # No plot for this test

            # Create sheet1 using create_sheet1_variables
            # For this test, we'll create a simple sheet1
            sheet1 = df.copy()
            sheet1['test_col'] = np.random.randn(len(df))

            # Convert reps to DataFrame
            import pandas as pd
            reps_df = pd.DataFrame([{
                'rep': r.rep,
                'amp_up_mm': r.amp_up_mm,
                'dur_up_s': r.dur_up_s,
                'dur_down_s': r.dur_down_s,
                'dur_seated_after_s': r.dur_seated_after_s,
                'ok_up': r.ok_up
            } for r in reps])

            # Create params dict
            params = {
                'mass_kg': 75.0,
                'dt': 0.01,
                'time_col': 'Tiempo'
            }

            export_to_excel(
                sheet1=sheet1,
                reps_df=reps_df,
                params=params,
                plot_path=plot_path,
                out_excel=excel_path
            )

            # Check file was created
            assert os.path.exists(excel_path)

            # Load and verify workbook
            wb = load_workbook(excel_path)
            sheet_names = wb.sheetnames

            assert 'Hoja1_Variables' in sheet_names
            assert 'Hoja2_Parametros_Grafica' in sheet_names
            assert 'Hoja3_Resultados_Repeticion' in sheet_names

    def test_sheet1_contains_data(self):
        """Sheet1 should contain the DataFrame data."""
        df, reps = self.create_test_data()

        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = os.path.join(temp_dir, 'test.xlsx')

            # Create sheet1
            sheet1 = df.copy()
            sheet1['test_col'] = np.random.randn(len(df))

            # Convert reps to DataFrame
            import pandas as pd
            reps_df = pd.DataFrame([{
                'rep': r.rep,
                'amp_up_mm': r.amp_up_mm,
                'dur_up_s': r.dur_up_s,
                'dur_down_s': r.dur_down_s,
                'dur_seated_after_s': r.dur_seated_after_s,
                'ok_up': r.ok_up
            } for r in reps])

            # Create params dict
            params = {
                'mass_kg': 75.0,
                'dt': 0.01,
                'time_col': 'Tiempo'
            }

            export_to_excel(
                sheet1=sheet1,
                reps_df=reps_df,
                params=params,
                plot_path=None,
                out_excel=excel_path
            )

            wb = load_workbook(excel_path)
            ws1 = wb['Hoja1_Variables']

            # Check headers
            assert ws1['A1'].value == 'Tiempo'
            assert ws1['B1'].value == 'BCM Z'
            assert ws1['C1'].value == 'Velocidad'

            # Check data exists
            assert ws1['A2'].value is not None
            assert ws1['B2'].value is not None

    def test_sheet3_contains_rep_results(self):
        """Sheet3 should contain repetition results."""
        df, reps = self.create_test_data()

        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = os.path.join(temp_dir, 'test.xlsx')

            # Create sheet1
            sheet1 = df.copy()
            sheet1['test_col'] = np.random.randn(len(df))

            # Convert reps to DataFrame
            import pandas as pd
            reps_df = pd.DataFrame([{
                'rep': r.rep,
                'amp_up_mm': r.amp_up_mm,
                'dur_up_s': r.dur_up_s,
                'dur_down_s': r.dur_down_s,
                'dur_seated_after_s': r.dur_seated_after_s,
                'ok_up': r.ok_up
            } for r in reps])

            # Create params dict
            params = {
                'mass_kg': 75.0,
                'dt': 0.01,
                'time_col': 'Tiempo'
            }

            export_to_excel(
                sheet1=sheet1,
                reps_df=reps_df,
                params=params,
                plot_path=None,
                out_excel=excel_path
            )

            wb = load_workbook(excel_path)
            ws3 = wb['Hoja3_Resultados_Repeticion']

            # Check headers
            assert ws3['A1'].value == 'rep'
            assert ws3['B1'].value == 'amp_up_mm'

            # Check data
            assert ws3['A2'].value == 1
            assert ws3['B2'].value == 50.0
            assert ws3['A3'].value == 2
            assert ws3['B3'].value == 48.0

    def test_handles_empty_reps(self):
        """Should handle empty repetitions list."""
        df, _ = self.create_test_data()

        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = os.path.join(temp_dir, 'test.xlsx')

            # Create sheet1
            sheet1 = df.copy()

            # Empty reps DataFrame
            import pandas as pd
            reps_df = pd.DataFrame()

            # Create params dict
            params = {
                'mass_kg': 75.0,
                'dt': 0.01,
                'time_col': 'Tiempo'
            }

            export_to_excel(
                sheet1=sheet1,
                reps_df=reps_df,
                params=params,
                plot_path=None,
                out_excel=excel_path
            )

            wb = load_workbook(excel_path)
            ws3 = wb['Hoja3_Resultados_Repeticion']

            # Should have message about no repetitions
            assert 'No se detectaron' in str(ws3['A1'].value)

    def test_handles_missing_plot(self):
        """Should handle missing plot file gracefully."""
        df, reps = self.create_test_data()

        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = os.path.join(temp_dir, 'test.xlsx')
            plot_path = os.path.join(temp_dir, 'nonexistent.png')

            # Create sheet1
            sheet1 = df.copy()

            # Convert reps to DataFrame
            import pandas as pd
            reps_df = pd.DataFrame([{
                'rep': r.rep,
                'amp_up_mm': r.amp_up_mm,
                'dur_up_s': r.dur_up_s,
                'dur_down_s': r.dur_down_s,
                'dur_seated_after_s': r.dur_seated_after_s,
                'ok_up': r.ok_up
            } for r in reps])

            # Create params dict
            params = {
                'mass_kg': 75.0,
                'dt': 0.01,
                'time_col': 'Tiempo'
            }

            # Should not crash
            export_to_excel(
                sheet1=sheet1,
                reps_df=reps_df,
                params=params,
                plot_path=plot_path,
                out_excel=excel_path
            )

            assert os.path.exists(excel_path)


class TestExportToJson:
    """Tests for export_to_json function."""

    def test_exports_json_file(self):
        """Should create JSON file with correct content."""
        params = {
            'mass_kg': 75.0,
            'window': 30,
            'n_reps': 5,
            'results': {
                'mean_duration': 1.5,
                'mean_velocity': 0.6,
                'total_work': 450.0
            }
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            json_path = os.path.join(temp_dir, 'test.json')

            export_to_json(params, json_path)

            assert os.path.exists(json_path)

            # Verify content
            with open(json_path, 'r', encoding='utf-8') as f:
                loaded = json.load(f)

            assert loaded['mass_kg'] == 75.0
            assert loaded['results']['mean_duration'] == 1.5

    def test_handles_special_characters(self):
        """Should handle special characters and unicode."""
        params = {
            'subject': 'Paciente_Ñoño_Álvarez',
            'notes': 'Evaluación post-operatoria: prótesis de cadera',
            'results': {'score': 85.5}
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            json_path = os.path.join(temp_dir, 'test.json')

            export_to_json(params, json_path)

            with open(json_path, 'r', encoding='utf-8') as f:
                loaded = json.load(f)

            assert loaded['subject'] == 'Paciente_Ñoño_Álvarez'
            assert 'post-operatoria' in loaded['notes']


class TestExportAdvancedSheet3:
    """Tests for export_advanced_sheet3 function."""

    def create_test_phase_df(self):
        """Create test DataFrame for advanced sheet export."""
        return pd.DataFrame({
            'phase': ['concéntrica', 'excéntrica', 'sentado'] * 2,
            'rep_idx': [1, 1, 1, 2, 2, 2],
            'duration_s': [0.8, 0.7, 1.2, 0.85, 0.75, 1.1],
            'BCM_Z_range': [0.15, 0.12, 0.05, 0.14, 0.11, 0.04],
            'BCM_Z_stats_mean': [0.08, 0.002, 0.085, 0.075, 0.07, 0.08],  # mean values
            'hip_metrics_mean': [0.12, 0.08, 0.14, 0.11, 0.09, 0.13],  # mean values
            'emg_peak': [0.45, 0.32, 0.12, 0.48, 0.35, 0.15],
            'emg_duration_ms': [450, 380, 120, 480, 420, 150],
            'cop_metrics_mean': [0.02, 0.015, 0.025, 0.018, 0.022, 0.019],  # mean values
            'work_J': [45.2, 38.1, 12.3, 42.8, 35.6, 11.8]
        })

    def test_exports_advanced_sheet(self):
        """Should create advanced sheet with correct data."""
        phase_df = self.create_test_phase_df()
        wb = Workbook()

        export_advanced_sheet3(wb, phase_df)

        assert 'Hoja3_Kinematic_&_Forces_like' in wb.sheetnames

        ws = wb['Hoja3_Kinematic_&_Forces_like']

        # Check headers
        assert ws['A1'].value == 'phase'
        assert ws['B1'].value == 'rep_idx'
        assert ws['C1'].value == 'duration_s'

        # Check data
        assert ws['A2'].value == 'concéntrica'
        assert ws['B2'].value == 1
        assert ws['C2'].value == 0.8

    def test_handles_empty_dataframe(self):
        """Should handle empty DataFrame gracefully."""
        phase_df = pd.DataFrame()
        wb = Workbook()

        export_advanced_sheet3(wb, phase_df)

        ws = wb['Hoja3_Kinematic_&_Forces_like']
        assert 'Sin repetición detectada' in str(ws['A1'].value)

    def test_handles_nan_values(self):
        """Should handle NaN values in advanced sheet."""
        phase_df = self.create_test_phase_df()
        phase_df.loc[0, 'work_J'] = np.nan
        phase_df.loc[1, 'emg_peak'] = np.nan

        wb = Workbook()
        export_advanced_sheet3(wb, phase_df)

        ws = wb['Hoja3_Kinematic_&_Forces_like']

        # Check that NaN values are handled (should be None in Excel)
        # work_J is in column J, emg_peak is in column G
        work_j_values = []
        emg_peak_values = []

        for row in range(2, len(phase_df) + 2):
            work_j_values.append(ws[f'J{row}'].value)  # work_J column
            emg_peak_values.append(ws[f'G{row}'].value)  # emg_peak column

        # Should have at least one None value from the NaN we set
        assert None in work_j_values or None in emg_peak_values


class TestExportIntegration:
    """Integration tests for export functionality."""

    def test_full_export_workflow(self):
        """Test complete export workflow."""
        # Create comprehensive test data
        df = pd.DataFrame({
            'Tiempo': np.arange(0, 2.0, 0.01),
            'BCM Z': 600 + 50 * np.sin(np.arange(0, 2.0, 0.01) * np.pi),
            'Velocidad': 0.5 * np.cos(np.arange(0, 2.0, 0.01) * np.pi),
            'Aceleración': -0.5 * np.sin(np.arange(0, 2.0, 0.01) * np.pi),
        })

        reps = [
            RepResult(
                rep=1,
                idx_conc_start=10,
                idx_peak=20,
                idx_ecc_end=30,
                t_conc_start_s=0.1,
                t_peak_s=0.2,
                t_ecc_end_s=0.3,
                amp_up_mm=140.0,
                amp_down_mm=135.0,
                dur_up_s=1.4,
                dur_down_s=1.3,
                dur_seated_after_s=0.5,
                vmax_up_m_s=0.7,
                vmin_down_m_s=-0.6,
                t_to_vmax_up_s=0.7,
                t_to_vmin_down_s=0.5,
                t_pos_acc_up_s=0.35,
                t_neg_acc_down_s=0.25,
                pmax_up_W=98.0,
                pmin_down_W=-85.0,
                work_up_J=68.6,
                work_down_J=55.0,
                ok_up=1
            ),
            RepResult(
                rep=2,
                idx_conc_start=40,
                idx_peak=50,
                idx_ecc_end=60,
                t_conc_start_s=0.4,
                t_peak_s=0.5,
                t_ecc_end_s=0.6,
                amp_up_mm=145.0,
                amp_down_mm=140.0,
                dur_up_s=1.5,
                dur_down_s=1.4,
                dur_seated_after_s=0.6,
                vmax_up_m_s=0.72,
                vmin_down_m_s=-0.65,
                t_to_vmax_up_s=0.72,
                t_to_vmin_down_s=0.52,
                t_pos_acc_up_s=0.36,
                t_neg_acc_down_s=0.26,
                pmax_up_W=102.0,
                pmin_down_W=-88.0,
                work_up_J=76.5,
                work_down_J=62.0,
                ok_up=1
            )
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = os.path.join(temp_dir, 'full_test.xlsx')
            json_path = os.path.join(temp_dir, 'params.json')

            # Export Excel
            # Create sheet1
            sheet1 = df.copy()
            sheet1['test_col'] = np.random.randn(len(df))

            # Convert reps to DataFrame
            reps_df = pd.DataFrame([{
                'rep': r.rep,
                'amp_up_mm': r.amp_up_mm,
                'dur_up_s': r.dur_up_s,
                'dur_down_s': r.dur_down_s,
                'dur_seated_after_s': r.dur_seated_after_s,
                'ok_up': r.ok_up
            } for r in reps])

            # Create params dict
            params = {
                'analysis_type': 'STS',
                'subject_mass': 75.0,
                'n_repetitions': len(reps),
                'total_duration': df['Tiempo'].max(),
                'mass_kg': 75.0,
                'dt': 0.01,
                'time_col': 'Tiempo'
            }

            export_to_excel(
                sheet1=sheet1,
                reps_df=reps_df,
                params=params,
                plot_path=None,
                out_excel=excel_path
            )

            # Export JSON
            params = {
                'analysis_type': 'STS',
                'subject_mass': 75.0,
                'n_repetitions': len(reps),
                'total_duration': df['Tiempo'].max()
            }
            export_to_json(params, json_path)

            # Verify files exist
            assert os.path.exists(excel_path)
            assert os.path.exists(json_path)

            # Verify Excel content
            wb = load_workbook(excel_path)
            assert len(wb.sheetnames) == 3

            # Verify JSON content
            with open(json_path, 'r') as f:
                loaded_params = json.load(f)
            assert loaded_params['subject_mass'] == 75.0
            assert loaded_params['n_repetitions'] == 2