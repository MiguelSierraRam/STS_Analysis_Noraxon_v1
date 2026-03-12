import os
import tempfile
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

from src.export import export_advanced_sheet3
from src.plotting import generate_plots
from src.metrics import RepResult
from src.metadata import read_metadata


def test_export_advanced_sheet3_nonempty():
    wb = Workbook()
    df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
    export_advanced_sheet3(wb, df)
    assert 'Hoja3_Kinematic_&_Forces_like' in wb.sheetnames
    ws = wb['Hoja3_Kinematic_&_Forces_like']
    # first row should contain headers
    assert ws.cell(row=1, column=1).value == 'A'
    assert ws.cell(row=1, column=2).value == 'B'
    # column width set to at least 14
    assert ws.column_dimensions['A'].width >= 14


def test_export_advanced_sheet3_empty():
    wb = Workbook()
    df = pd.DataFrame()
    export_advanced_sheet3(wb, df)
    ws = wb['Hoja3_Kinematic_&_Forces_like']
    assert ws['A1'].value == 'Sin repetición detectada'


def test_generate_plots_basic(tmp_path):
    t = np.linspace(0, 1, 11)
    z = np.sin(2 * np.pi * t)
    v = np.gradient(z, t)
    idx_starts = [0]
    idx_peaks = [5]
    idx_ends = [10]
    # use compute_metrics helper to create a valid RepResult
    from src.metrics import compute_metrics
    # build minimal arrays for metrics
    n = len(t)
    z_mm = z
    vel = v
    acc = np.gradient(v, t)
    power = np.zeros_like(t)
    dt = float(np.median(np.diff(t)))
    rep = compute_metrics(1, 0, 5, 10, n - 1, n, z_mm, vel, acc, power, t, dt)

    out_prefix = str(tmp_path / 'out')
    plot_path, rep_dir = generate_plots(t, z, v, idx_starts, idx_peaks, idx_ends, [rep], out_prefix)
    assert os.path.isfile(plot_path)
    # per-rep directory should exist and contain at least one png
    assert os.path.isdir(rep_dir)
    files = os.listdir(rep_dir)
    assert any(f.endswith('.png') for f in files)


def test_read_metadata(tmp_path):
    # build a simple excel file
    file_path = tmp_path / 'meta.xlsx'
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # create sheet with appropriate rows (initial blank row to align indices)
        meta = pd.DataFrame([
            ['', None],
            ['', 'Codigo1'],
            ['', 180],
            ['', 75],
            ['', 'TestType'],
            ['', '2025-01-01'],
            ['', 45],
        ])
        meta.to_excel(writer, sheet_name='MetaData_&_Parameters', index=False, header=False)
    result = read_metadata(str(file_path))
    assert result['Codigo'] == 'Codigo1'
    assert result['Altura'] == 180
    assert result['Peso_kg'] == 75
    assert result['Test'] == 'TestType'
    assert result['Fecha_Test'] == '2025-01-01'
    assert result['Altura_silla'] == 45
