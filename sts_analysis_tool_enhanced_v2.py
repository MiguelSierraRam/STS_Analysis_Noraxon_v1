#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Herramienta de análisis STS (Sit-to-Stand) - Versión Enhanced v2.1

Features:
- Detección exacta de fases estilo Excel
- Soporte para Hip Z (desplazamiento y velocidad)
- Soporte para EMG (electromiografía)
- Soporte para CoP SD (estabilidad Centro de Presión)
- "Acc Phases" (marcadores acumulativos de fase)
- Metadatos Noraxon (código, altura, peso, etc.)
- Hoja3 ampliada (Kinematic & Forces like)
- Batch processing + CSV export
- Logging centralizado profesional
- Configuración vía config.yaml

Requiere: numpy, pandas, matplotlib, openpyxl, pyyaml

Modo individual:
    python sts_analysis_tool_enhanced_v2.py --input data/input/file.xlsx --mass-kg 75

Modo batch:
    python sts_analysis_tool_enhanced_v2.py --batch data/input --out data/output --mass-kg 75 --csv

Con configuración custom:
    python sts_analysis_tool_enhanced_v2.py --config custom_config.yaml --input file.xlsx
"""

import argparse
import os
import json
from typing import Optional, Tuple, List
import logging

import numpy as np
import pandas as pd

from src.utils import centered_slope, cumulative_trapezoid, detect_column
from src.detection import (
    compute_vector_displacements,
    compute_windows,
    detect_conc_starts,
    detect_peaks,
    detect_ecc_ends,
    pair_repetitions,
    compute_phase_events,
    compute_acc_phases,
)
from src.metrics import RepResult, compute_metrics, mark_ok_repetitions
from src.plotting import generate_plots
from src.export import create_sheet1_variables, export_to_excel, export_to_json, export_advanced_sheet3
from src.advanced_metrics import PhaseComputer, detect_emg_columns, detect_cop_columns
from src.metadata import read_metadata
from src.config import load_config, get_config
from src.logger import LoggerManager, get_logger

# Importar para Excel
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# Inicializar logger
logger = get_logger(__name__)


def load_participants_db(db_path: str = 'data/participants_db.xlsx') -> pd.DataFrame:
    """Carga la base de datos de participantes desde Excel."""
    if not os.path.exists(db_path):
        logger.warning(f"BD de participantes no encontrada: {db_path}. Usando masa por defecto.")
        return pd.DataFrame()
    try:
        df = pd.read_excel(db_path, engine='openpyxl')
        logger.info(f"BD de participantes cargada: {len(df)} registros")
        return df
    except Exception as e:
        logger.error(f"Error cargando BD de participantes: {e}")
        return pd.DataFrame()


def get_participant_mass(participant_id: str, db_df: pd.DataFrame) -> Optional[float]:
    """Obtiene la masa del participante desde la BD."""
    if db_df.empty:
        return None
    # Asumir columna 'participant_id' y 'mass_kg'
    row = db_df[db_df['participant_id'].astype(str).str.upper() == participant_id.upper()]
    if not row.empty:
        return float(row['mass_kg'].iloc[0])
    return None


def export_processed_data(input_file: str, participant_id: str, mass_kg: float, out_dir: str, hoja3_df: pd.DataFrame = None):
    """Exporta datos procesados a BD."""
    db_path = os.path.join(out_dir, 'processed_data.csv')
    
    # Crear datos resumidos
    summary = {
        'participant_id': participant_id,
        'input_file': os.path.basename(input_file),
        'mass_kg': mass_kg,
        'processed_at': pd.Timestamp.now().isoformat(),
        'n_reps': len(hoja3_df['Repeticion'].unique()) if hoja3_df is not None else 0,
    }
    
    # Agregar métricas promedio si hay datos
    if hoja3_df is not None and not hoja3_df.empty:
        # Promedios por fase
        for phase in ['Concéntrica', 'Excéntrica', 'Sentado']:
            phase_data = hoja3_df[hoja3_df['Fase'] == phase]
            if not phase_data.empty:
                # Calcular promedios de las columnas disponibles
                if 'BCM_Z_Power_Mean (W)' in phase_data.columns and phase_data['BCM_Z_Power_Mean (W)'].notna().any():
                    summary[f'{phase.lower()}_power_mean_avg'] = phase_data['BCM_Z_Power_Mean (W)'].mean()
                if 'BCM_Z_Work (J)' in phase_data.columns and phase_data['BCM_Z_Work (J)'].notna().any():
                    summary[f'{phase.lower()}_work_avg'] = phase_data['BCM_Z_Work (J)'].mean()
                if 'BCM_Z_Vel_Max (m/s)' in phase_data.columns:
                    summary[f'{phase.lower()}_vel_max_avg'] = phase_data['BCM_Z_Vel_Max (m/s)'].mean()
                if 'BCM_Z_Range (m)' in phase_data.columns:
                    summary[f'{phase.lower()}_range_avg'] = phase_data['BCM_Z_Range (m)'].mean()
                if 'BCM_Z_Vel_Range (m/s)' in phase_data.columns:
                    summary[f'{phase.lower()}_vel_range_avg'] = phase_data['BCM_Z_Vel_Range (m/s)'].mean()
                if 'BCM_Z_Acc_Mean (m/s²)' in phase_data.columns:
                    summary[f'{phase.lower()}_acc_mean_avg'] = phase_data['BCM_Z_Acc_Mean (m/s²)'].mean()
                # EMG promedio de max por fase
                emg_max_cols = [c for c in phase_data.columns if 'EMG' in c and 'Max' in c]
                if emg_max_cols:
                    summary[f'{phase.lower()}_emg_max_avg'] = phase_data[emg_max_cols].mean().mean()
    
    # Cargar BD existente o crear nueva
    if os.path.exists(db_path):
        existing_df = pd.read_csv(db_path)
        new_df = pd.DataFrame([summary])
        updated_df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        updated_df = pd.DataFrame([summary])
    
    # Guardar
    updated_df.to_csv(db_path, index=False)
    logger.info(f"BD procesada actualizada: {db_path}")


def run_tool_enhanced(
    file_path: str,
    sheet_name: str = 'Reducido',
    time_col: str = 'time',
    disp_col: Optional[str] = None,
    mass_kg: Optional[float] = None,
    window: int = 30,
    n_positive: int = 30,
    vel_th_m_s: float = 0.1,
    ok_th: float = 0.85,
    half_window_derivative: int = 3,
    out_prefix: Optional[str] = None,
    out_dir: Optional[str] = None,
    csv_export: bool = False,
    make_plot: bool = True,
    per_rep_plots: bool = True,
    close_last_seated_at_end: bool = True,
    participant_id: Optional[str] = None,
    participants_db: Optional[pd.DataFrame] = None,
) -> Tuple[str, Optional[str], str, Optional[str]]:
    """Ejecuta análisis STS completo con funcionalidades avanzadas."""
    logger.info(f"Iniciando análisis: {file_path}")
    logger.debug(f"  Hoja: {sheet_name}, Tiempo col: {time_col}")
    
    # -------- 1. LEER DATOS --------
    logger.debug("Leyendo datos de Excel...")
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    if time_col not in df.columns:
        raise ValueError(f'No se encontró columna "{time_col}".')
    
    t = df[time_col].astype(float).to_numpy()
    if len(t) < 3:
        raise ValueError('No hay suficientes muestras.')
    
    logger.debug(f"Datos cargados: {len(t)} muestras")
    
    dt = float(pd.Series(np.diff(t)).median())
    n = len(t)
    
    # Detectar desplazamiento BCM Z
    if disp_col is None:
        disp_col = detect_column(
            df,
            candidates=['Disp_BCM_Z', 'BCM_Z', 'Body center of mass-z (mm)'],
            contains_all=['body center of mass', 'z', '(mm)']
        )
        if disp_col is None:
            for c in df.columns:
                cl = str(c).lower()
                if ('body center of mass' in cl) and ('z' in cl) and ('mm' in cl):
                    disp_col = c
                    break
    
    if disp_col is None:
        raise ValueError('No se encontró columna BCM Z.')
    
    z_mm = df[disp_col].astype(float).to_numpy()
    z_m = z_mm / 1000.0
    
    # -------- 2. CALCULAR VARIABLES DERIVADAS --------
    vel_m_s = centered_slope(z_m, dt, half_window_derivative)
    acc_m_s2 = centered_slope(vel_m_s, dt, half_window_derivative)
    
    force_est_N = np.full_like(vel_m_s, np.nan)
    if mass_kg is not None:
        force_est_N[:] = mass_kg * 9.80665
        power_W = force_est_N * vel_m_s
    else:
        power_W = np.full_like(vel_m_s, np.nan)
    
    work_cum_J = cumulative_trapezoid(power_W, dt)
    
    # -------- 3. DETECCIÓN DE EVENTOS --------
    vector_disp = compute_vector_displacements(z_mm)
    future_sum, previous_sum = compute_windows(vector_disp, window)
    
    vel_conc_flag = np.where(vel_m_s > vel_th_m_s, 1, np.nan)
    vel_ecc_flag = np.where(vel_m_s < -vel_th_m_s, 1, np.nan)
    
    conc_start = detect_conc_starts(future_sum, vel_conc_flag, n_positive)
    conc_exc = detect_peaks(z_mm, window)
    ecc_end = detect_ecc_ends(previous_sum, vel_ecc_flag, n_positive)
    
    conc_event, conc_graph, ecc_event, ecc_graph, any_phase_event = compute_phase_events(
        conc_start, conc_exc, ecc_end
    )
    
    # -------- 4. EMPAREJAR REPETICIONES --------
    idx_starts = list(np.where(conc_start == 10000)[0])
    idx_peaks = list(np.where(conc_exc == 10000)[0])
    idx_ends = list(np.where(ecc_end == 10000)[0])
    
    reps_data, phase_id, phase_label = pair_repetitions(idx_starts, idx_peaks, idx_ends, n)
    
    # -------- 5. CALCULAR ACC_PHASES Y MARKER_TO_TIME --------
    acc_phases, marker_to_time, rep_events = compute_acc_phases(conc_start, conc_exc, ecc_end, t)
    
    # -------- 6. CALCULAR MÉTRICAS POR REPETICIÓN --------
    reps: List[RepResult] = []
    for rep_num, (s, p, e) in enumerate(reps_data, start=1):
        next_s = idx_starts[rep_num] if rep_num < len(idx_starts) else n - 1
        rep = compute_metrics(rep_num, s, p, e, next_s, n, z_mm, vel_m_s, acc_m_s2, power_W, t, dt)
        reps.append(rep)
    
    mark_ok_repetitions(reps, ok_th)
    
    # -------- 7. CREAR HOJA1 --------
    rep_id = np.full(n, np.nan)
    for rr in reps:
        s, p, e = rr.idx_conc_start, rr.idx_peak, rr.idx_ecc_end
        if s is not None and p is not None:
            rep_id[s:p+1] = rr.rep
        if p is not None and e is not None:
            rep_id[p:e+1] = rr.rep
        if e is not None:
            next_starts = [x for x in idx_starts if x > e]
            ns = next_starts[0] if next_starts else n
            rep_id[e+1:ns] = rr.rep
    
    sheet1 = create_sheet1_variables(
        df, z_mm, vel_m_s, acc_m_s2, force_est_N, power_W, work_cum_J,
        vector_disp, future_sum, previous_sum,
        vel_conc_flag, vel_ecc_flag,
        conc_start, conc_exc, ecc_end,
        conc_event, conc_graph, ecc_event, ecc_graph, any_phase_event,
        phase_id, phase_label, rep_id
    )
    
    sheet1['Acc Phases'] = acc_phases
    
    # -------- 8. CREAR HOJA3 AMPLIADA --------
    metadata = read_metadata(file_path, mass_kg, participant_id, participants_db)
    pc = PhaseComputer(df, time_col)
    
    # Agregar columnas calculadas al df para PhaseComputer
    df_temp = df.copy()
    df_temp['BCM_Z_Vel (m/s)'] = vel_m_s
    df_temp['BCM_Z_Power (W)'] = power_W
    pc_temp = PhaseComputer(df_temp, time_col)
    
    emg_cols = detect_emg_columns(df)
    cop_cols = detect_cop_columns(df)
    
    disp_hipz = next((c for c in df.columns if 'hip rt-z' in str(c).lower() 
                     or ('hip' in str(c).lower() and 'z' in str(c).lower())), None)
    
    hoja3_rows = []
    for rep_event in rep_events:
        r = rep_event.rep
        
        for phase_name in ['Concéntrica', 'Excéntrica', 'Sentado']:
            m1 = (r - 1) * 3 + 1
            m2 = m1 + 1
            m3 = m1 + 2
            
            if phase_name == 'Concéntrica':
                sm, em = m1, m2
            elif phase_name == 'Excéntrica':
                sm, em = m2, m3
            else:
                sm, em = m3, m3 + 1
            
            if sm not in marker_to_time:
                continue
            
            t0 = marker_to_time[sm]
            t1 = marker_to_time[em] if em in marker_to_time else (float(t[-1]) if close_last_seated_at_end else t0)
            
            row = {
                'Codigo': metadata['Codigo'],
                'Altura': metadata['Altura'],
                'Peso_kg': metadata['Peso_kg'],
                'Altura_silla': metadata['Altura_silla'],
                'Test': metadata['Test'],
                'Fecha_Test': metadata['Fecha_Test'],
                'Repeticion': r,
                'Fase': phase_name,
                'Inicio_Tiempo': t0,
                'Final_Tiempo': t1,
                'Duracion_Fase (s)': t1 - t0,  # Nueva: duración de la fase
            }
            
            # BCM Z
            mx, mn, rg = pc.range_mm_to_m(disp_col, t0, t1)
            row['BCM_Z_Max (m)'] = mx
            row['BCM_Z_Min (m)'] = mn
            row['BCM_Z_Range (m)'] = rg
            
            mV, xV, minV = pc.stats(None, vel_m_s, t0, t1)
            row['BCM_Z_Vel_Mean (m/s)'] = mV
            row['BCM_Z_Vel_Max (m/s)'] = xV
            
            # BCM Z Power and Work with interpolation
            if metadata['Peso_kg'] is not None:
                # Interpolate velocity over the time period
                t_interp = np.linspace(t0, t1, num=100)  # 100 points for interpolation
                vel_interp = np.interp(t_interp, t, vel_m_s)
                p_interp = metadata['Peso_kg'] * 9.80665 * vel_interp
                work_calc = np.trapezoid(p_interp, t_interp)
                mean_p_calc = np.mean(p_interp)
                max_p_calc = np.max(p_interp)
                
                row['BCM_Z_Power_Mean (W)'] = mean_p_calc
                row['BCM_Z_Power_Max (W)'] = max_p_calc  # Nueva variable: potencia máxima
                row['BCM_Z_Work (J)'] = work_calc
            else:
                row['BCM_Z_Power_Mean (W)'] = None
                row['BCM_Z_Power_Max (W)'] = None
                row['BCM_Z_Work (J)'] = None
            
            # Calcular BCM_Z_Vel_Range (m/s): rango de velocidad
            vel_slice = vel_m_s[(t >= t0) & (t <= t1)]
            if len(vel_slice) > 0:
                row['BCM_Z_Vel_Range (m/s)'] = np.ptp(vel_slice)  # peak-to-peak
            else:
                row['BCM_Z_Vel_Range (m/s)'] = None
            
            # Jerk (derivada de aceleración)
            jerk_m_s3 = centered_slope(acc_m_s2, dt, half_window_derivative)
            jerk_slice = jerk_m_s3[(t >= t0) & (t <= t1)]
            if len(jerk_slice) > 0:
                row['BCM_Z_Jerk_Mean (m/s³)'] = np.mean(jerk_slice)
                row['BCM_Z_Jerk_Max (m/s³)'] = np.max(jerk_slice)
            
            # Ratios
            if 'BCM_Z_Power_Mean (W)' in row and row['BCM_Z_Power_Mean (W)'] is not None and metadata['Peso_kg']:
                row['Power_to_Weight_Ratio'] = row['BCM_Z_Power_Mean (W)'] / (metadata['Peso_kg'] * 9.80665)  # W/N
            if 'BCM_Z_Vel_Max (m/s)' in row and row['BCM_Z_Vel_Max (m/s)'] is not None:
                row['Vel_to_Height_Ratio'] = row['BCM_Z_Vel_Max (m/s)'] / metadata.get('Altura', 1.7)  # m/s / m
            
            # Hip Z adicionales
            if disp_hipz is not None:
                hmx, hmn, hrg = pc.range_mm_to_m(disp_hipz, t0, t1)
                row['HIP_Z_Max (m)'] = hmx
                row['HIP_Z_Min (m)'] = hmn  # Nueva
                row['HIP_Z_Range (m)'] = hrg
                
                # Velocidad de Hip Z (si hay columna de vel hip)
                hip_vel_col = next((c for c in df.columns if 'hip' in str(c).lower() and 'vel' in str(c).lower()), None)
                if hip_vel_col:
                    hip_vel_mean, hip_vel_max, _ = pc.stats(hip_vel_col, None, t0, t1)
                    row['HIP_Z_Vel_Mean (m/s)'] = hip_vel_mean  # Nueva
                    row['HIP_Z_Vel_Max (m/s)'] = hip_vel_max  # Nueva
            
            # EMG variables adicionales (máximos y mínimos)
            for emg in emg_cols:
                emg_mean, emg_max, emg_min = pc.stats(emg, None, t0, t1)
                row[f'EMG_{emg}'] = emg_mean  # Media (ya existente)
                row[f'EMG_{emg}_Max (%)'] = emg_max  # Nueva: máximo
                row[f'EMG_{emg}_Min (%)'] = emg_min  # Nueva: mínimo
            
            # CoP variables adicionales (máximos y mínimos)
            for key, col in cop_cols.items():
                if col:
                    cop_mean = pc.mean(col, t0, t1)
                    row[f'CoP_{key}'] = cop_mean  # Ya existente
                    
                    # Nuevas: máximo y mínimo de CoP
                    cop_max = pc.stats(col, None, t0, t1)[1]
                    cop_min = pc.stats(col, None, t0, t1)[2]
                    row[f'CoP_{key}_Max'] = cop_max
                    row[f'CoP_{key}_Min'] = cop_min
            
            hoja3_rows.append(row)
    
    hoja3_df = pd.DataFrame(hoja3_rows)
    
    # -------- 9. CONFIGURAR RUTAS DE SALIDA --------
    if not out_prefix:
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            out_prefix = os.path.join(out_dir, base_name + '_analysis_enhanced')
        else:
            out_prefix = os.path.splitext(file_path)[0] + '_analysis_enhanced'
    
    plot_path, rep_dir = generate_plots(
        t, z_mm, vel_m_s,
        idx_starts, idx_peaks, idx_ends,
        reps, out_prefix,
        make_plot, per_rep_plots
    )
    
    # -------- 10. EXPORTAR DATOS --------
    params = {
        'input_file': file_path,
        'sheet_name': sheet_name,
        'mass_kg': mass_kg,
        'window': window,
        'n_positive': n_positive,
        'vel_th_m_s': vel_th_m_s,
        'ok_th': ok_th,
        'n_reps': len(reps),
    }
    
    # Excel
    out_excel = out_prefix + '.xlsx'
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Hoja1_Variables'
    
    for j, col in enumerate(sheet1.columns, start=1):
        cell = ws1.cell(row=1, column=j, value=str(col))
        cell.font = Font(bold=True)
    for i, (_, row) in enumerate(sheet1.iterrows(), start=2):
        for j, val in enumerate(row, start=1):
            if isinstance(val, (np.floating, float)) and (np.isnan(val) or np.isinf(val)):
                val = None
            ws1.cell(row=i, column=j, value=val)
    
    for j in range(1, len(sheet1.columns)+1):
        ws1.column_dimensions[get_column_letter(j)].width = min(max(12, 14), 28)
    
    # Hoja2
    ws2 = wb.create_sheet('Hoja2_Parametros_Grafica')
    for r, (k, v) in enumerate([('Archivo', params['input_file']), ('Hoja', params['sheet_name']), 
                                  ('Masa kg', params['mass_kg']), ('Reps', params['n_reps'])], start=1):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=v)
    
    if plot_path and os.path.exists(plot_path):
        img = XLImage(plot_path)
        img.width = 1100
        img.height = 550
        ws2.add_image(img, 'D2')
    
    # Hoja3
    export_advanced_sheet3(wb, hoja3_df)
    
    wb.save(out_excel)
    
    # JSON
    out_json = out_prefix + '_params.json'
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump(params, f, ensure_ascii=False, indent=2)
    
    # CSV
    out_csv = None
    if csv_export:
        try:
            out_csv = out_prefix + '_hoja3.csv'
            hoja3_df.to_csv(out_csv, index=False)
            logger.debug(f"CSV exportado: {out_csv}")
        except Exception as e:
            logger.warning(f"No se pudo exportar CSV: {e}")
    
    # Exportar a BD procesada si participant_id disponible
    if participant_id:
        export_processed_data(file_path, participant_id, mass_kg or 0, out_dir or os.path.dirname(out_prefix), hoja3_df)
    
    logger.info(f"✓ Análisis completado correctamente")
    logger.debug(f"  Excel: {out_excel}")
    logger.debug(f"  Plot: {plot_path}")
    logger.debug(f"  JSON: {out_json}")
    if out_csv:
        logger.debug(f"  CSV: {out_csv}")
    
    return out_excel, plot_path, out_json, out_csv


def analyze_file(
    input_path: str,
    out_dir: str = 'data/output',
    **kwargs
) -> Tuple[str, Optional[str], str, Optional[str]]:
    """Wrapper para procesar archivo con directorio de salida."""
    os.makedirs(out_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    out_prefix = os.path.join(out_dir, base_name + '_analysis_enhanced')
    
    return run_tool_enhanced(
        file_path=input_path,
        out_prefix=out_prefix,
        **kwargs
    )


def process_batch(
    batch_dir: str,
    out_dir: str = 'data/output',
    participants_db: pd.DataFrame = None,
    **kwargs
) -> List[Tuple[str, Tuple]]:
    """Procesa batch de archivos Excel."""
    os.makedirs(out_dir, exist_ok=True)
    entries = sorted([f for f in os.listdir(batch_dir) if f.lower().endswith('.xlsx')])
    results = []
    
    for name in entries:
        in_path = os.path.join(batch_dir, name)
        base = os.path.splitext(name)[0]
        out_prefix = os.path.join(out_dir, base + '_analysis_enhanced')
        out_xlsx = out_prefix + '.xlsx'
        
        # Extraer participant_id
        parts = name.split('_')
        participant_id = parts[1] if len(parts) > 1 else base
        mass_kg = kwargs.get('mass_kg')
        if mass_kg is None and participants_db is not None:
            mass_kg = get_participant_mass(participant_id, participants_db)
            if mass_kg is not None:
                logger.info(f"Masa para {participant_id}: {mass_kg} kg")
        
        # Skip si existe y es más nuevo (a menos que overwrite=True)
        if not kwargs.get('overwrite', False) and os.path.exists(out_xlsx):
            try:
                in_mtime = os.path.getmtime(in_path)
                out_mtime = os.path.getmtime(out_xlsx)
                if out_mtime >= in_mtime:
                    logger.info(f"SKIP (up-to-date): {name}")
                    continue
            except Exception:
                pass
        
        logger.info(f"Processing: {name}")
        try:
            kwargs_copy = kwargs.copy()
            kwargs_copy['mass_kg'] = mass_kg
            out = run_tool_enhanced(
                file_path=in_path,
                out_prefix=out_prefix,
                participant_id=participant_id,
                **kwargs_copy
            )
            results.append((name, out))
        except Exception as e:
            logger.error(f"ERROR procesando {name}: {e}", exc_info=True)
            results.append((name, (None, None, None, None)))
    
    return results


def main():
    """Interfaz principal con soporte batch, logging y configuración."""
    parser = argparse.ArgumentParser(
        description='STS Enhanced v2.1: Hip Z + EMG + CoP_SD + Batch + CSV + Logging + Config',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  # Archivo individual con config por defecto
  python sts_analysis_tool_enhanced_v2.py --input data/input/file.xlsx --mass-kg 75
  
  # Batch processing
  python sts_analysis_tool_enhanced_v2.py --batch data/input --out data/output --mass-kg 75 --csv
  
  # Con configuración custom
  python sts_analysis_tool_enhanced_v2.py --config custom.yaml --input file.xlsx
  
  # Ver configuración por defecto
  python sts_analysis_tool_enhanced_v2.py --show-config
        """
    )
    
    # Configuración y logging
    parser.add_argument('--config', default=None, help='Archivo config.yaml custom (default: config.yaml)')
    parser.add_argument('--show-config', action='store_true', help='Mostrar configuración y salir')
    parser.add_argument('--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'], 
                        default='INFO', help='Nivel de logging')
    parser.add_argument('--log-dir', default=None, help='Directorio custom para logs')
    
    # --input XOR --batch
    g = parser.add_mutually_exclusive_group(required=False)
    g.add_argument('--input', help='Archivo .xlsx individual')
    g.add_argument('--batch', help='Carpeta con .xlsx')
    
    # Parámetros comunes (con defaults de config.yaml)
    parser.add_argument('--out', default=None, help='Directorio de salida')
    parser.add_argument('--sheet', default=None, help='Hoja a procesar')
    parser.add_argument('--time-col', default=None, help='Columna de tiempo')
    parser.add_argument('--disp-col', default=None, help='Columna BCM Z')
    parser.add_argument('--mass-kg', type=float, default=None, help='Masa (kg)')
    parser.add_argument('--window', type=int, default=None, help='Ventana detección')
    parser.add_argument('--n-positive', type=int, default=None, help='N celdas positivas')
    parser.add_argument('--vel-th', type=float, default=None, help='Umbral vel (m/s)')
    parser.add_argument('--ok-th', type=float, default=None, help='Umbral amplitud')
    parser.add_argument('--csv', action='store_true', help='Exportar Hoja3 a CSV')
    parser.add_argument('--no-plot', action='store_true', help='Sin gráfico general')
    parser.add_argument('--no-per-rep', action='store_true', help='Sin gráficos por rep')
    parser.add_argument('--overwrite', action='store_true', help='Reprocesar todo (batch)')
    
    args = parser.parse_args()
    
    # ========== CARGAR CONFIGURACIÓN ==========
    # Cargar config.yaml
    config = load_config(args.config)
    
    # Cargar BD de participantes
    participants_db = load_participants_db()
    
    # Configurar logging
    LoggerManager.configure(log_dir=args.log_dir)
    logger.info(f"STS Analysis Tool Enhanced v2.1 iniciado")
    logger.debug(f"Archivo config: {args.config or 'config.yaml (default)'}")
    
    # Mostrar configuración si se solicita
    if args.show_config:
        print("\n" + "=" * 70)
        print("CONFIGURACIÓN ACTUAL")
        print("=" * 70)
        import yaml
        print(yaml.dump(config.to_dict(), default_flow_style=False, allow_unicode=True))
        return
    
    # ========== VALIDACIONES ==========
    if not args.input and not args.batch:
        parser.error("Se requiere --input o --batch")
    
    # ========== USAR DEFAULTS DE CONFIG.YAML PARA VALORES NO ESPECIFICADOS ==========
    sheet_name = args.sheet or config.get('sheet_name', 'Reducido')
    time_col = args.time_col or config.get('time_col', 'time')
    disp_col = args.disp_col or config.get('disp_col', None)
    out_dir = args.out or config.get('output_directory', 'data/output')
    window = args.window or config.get('window', 30)
    n_positive = args.n_positive or config.get('n_positive', 30)
    vel_th_m_s = args.vel_th or config.get('vel_th_m_s', 0.1)
    ok_th = args.ok_th or config.get('ok_th', 0.85)
    mass_kg = args.mass_kg or config.get('mass_kg', None)
    csv_export = args.csv or config.get('csv_export', False)
    make_plot = not args.no_plot and config.get('make_plot', True)
    per_rep_plots = not args.no_per_rep and config.get('per_rep_plots', True)
    
    kwargs = {
        'sheet_name': sheet_name,
        'time_col': time_col,
        'disp_col': disp_col,
        'mass_kg': mass_kg,
        'window': window,
        'n_positive': n_positive,
        'vel_th_m_s': vel_th_m_s,
        'ok_th': ok_th,
        'csv_export': csv_export,
        'make_plot': make_plot,
        'per_rep_plots': per_rep_plots,
    }
    
    logger.info(f"Parámetros de análisis:")
    logger.info(f"  - window: {window}")
    logger.info(f"  - n_positive: {n_positive}")
    logger.info(f"  - vel_th_m_s: {vel_th_m_s}")
    logger.info(f"  - ok_th: {ok_th}")
    logger.info(f"  - mass_kg: {mass_kg}")
    
    # ========== PROCESAMIENTO ==========
    if args.input:
        logger.info(f"Procesando archivo individual: {args.input}")
        
        # Extraer participant_id del nombre del archivo (ej. 2025-11-04-12-24_SMP011_D1_10STS_13REP_1.xlsx -> SMP011)
        base_name = os.path.basename(args.input)
        parts = base_name.split('_')
        participant_id = parts[1] if len(parts) > 1 else parts[0].split('.')[0]
        logger.info(f"Participant ID extraído: {participant_id}")
        
        # Usar masa de BD si no se especificó
        if mass_kg is None:
            mass_kg = get_participant_mass(participant_id, participants_db)
            if mass_kg is not None:
                logger.info(f"Masa obtenida de BD: {mass_kg} kg")
            else:
                logger.warning(f"No se encontró masa para {participant_id} en BD. Especifica --mass-kg")
        
        try:
            kwargs_copy = kwargs.copy()
            kwargs_copy['mass_kg'] = mass_kg  # Sobrescribir con masa de BD si aplica
            excel, plot, json_file, csv_file = run_tool_enhanced(
                file_path=args.input,
                out_dir=out_dir,
                participant_id=participant_id,
                participants_db=participants_db,
                **kwargs_copy
            )
            logger.info('✓ ANÁLISIS COMPLETADO')
            logger.info(f'  Excel:  {excel}')
            logger.info(f'  Plot:   {plot}')
            logger.info(f'  JSON:   {json_file}')
            if csv_file:
                logger.info(f'  CSV:    {csv_file}')
            
            # Mostrar en consola también
            print('\n✓ ANALYSIS COMPLETE')
            print(f'  Excel:  {excel}')
            print(f'  Plot:   {plot}')
            print(f'  JSON:   {json_file}')
            if csv_file:
                print(f'  CSV:    {csv_file}')
        except Exception as e:
            logger.error(f'ERROR: {e}', exc_info=True)
            print(f'\n✗ ERROR: {e}')
    else:
        logger.info(f"Iniciando batch processing: {args.batch} → {out_dir}")
        results = process_batch(
            batch_dir=args.batch,
            out_dir=out_dir,
            overwrite=args.overwrite,
            participants_db=participants_db,
            **kwargs
        )
        success = sum(1 for _, (e, _, _, _) in results if e is not None)
        failed = len(results) - success
        
        logger.info(f'✓ BATCH COMPLETADO')
        logger.info(f'  Total: {len(results)}, Éxito: {success}, Fallos: {failed}')
        
        # Mostrar en consola también
        print(f'\n✓ BATCH COMPLETE')
        print(f'  Total: {len(results)}')
        print(f'  Success: {success}, Failed: {failed}')


if __name__ == '__main__':
    main()
